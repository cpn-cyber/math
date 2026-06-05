"""Step 8: five-level grading and final Problem 1 visualization outputs."""

from __future__ import annotations

import os
from pathlib import Path
import sys
from typing import Any

os.environ.setdefault("OMP_NUM_THREADS", "1")
os.environ.setdefault("LOKY_MAX_CPU_COUNT", "1")

import numpy as np
import pandas as pd
import re

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    import yaml
except ImportError:  # pragma: no cover
    yaml = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from modules.grade_classifier import (  # noqa: E402
    GRADE_ORDER,
    classify_with_jenks,
    classify_with_kmeans,
    compare_grade_methods,
    find_boundary_nearby_papers,
    merge_classifications,
    read_final_scores,
    setup_grade_logger,
)
from modules.visualization import create_problem1_visualizations  # noqa: E402


def load_config(config_path: Path) -> dict[str, Any]:
    """Load YAML config if available; Step 8 also has safe defaults."""
    if yaml is None or not config_path.exists():
        return {}
    try:
        with config_path.open("r", encoding="utf-8") as file:
            return yaml.safe_load(file) or {}
    except Exception:
        return {}


def resolve_project_path(relative_path: str) -> Path:
    """Resolve a project-relative path."""
    return PROJECT_ROOT / relative_path


def _autosize_workbook(path: Path) -> None:
    """Apply light formatting to generated workbooks."""
    if load_workbook is None:
        return
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = 0
            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 36)
    workbook.save(path)


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> Path:
    """Write an Excel workbook with multiple sheets."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            safe_name = sheet_name[:31]
            frame.to_excel(writer, sheet_name=safe_name, index=False)
    _autosize_workbook(path)
    return path


def read_feature_table(path: Path) -> pd.DataFrame:
    """Read normalized features."""
    features = pd.read_excel(path)
    features = features.copy()
    features["paper_id"] = features["paper_id"].apply(
        lambda value: re.search(r"(\d+)", str(value)).group(1).zfill(2)
        if re.search(r"(\d+)", str(value))
        else str(value).zfill(2)
    )
    return features


def read_weight_table(path: Path) -> pd.DataFrame:
    """Read combined weights."""
    try:
        weights = pd.read_excel(path, sheet_name="combined_weights")
    except ValueError:
        weights = pd.read_excel(path)
    required = {"criterion", "indicator", "combined_weight"}
    missing = required - set(weights.columns)
    if missing:
        raise ValueError(f"权重表缺少必要字段: {sorted(missing)}")
    return weights


def make_distribution_table(
    kmeans_summary: pd.DataFrame,
    jenks_summary: pd.DataFrame,
    silhouette: float,
    consistency: float,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Create grade distribution and method summary tables."""
    distribution = kmeans_summary[["grade", "count", "score_min", "score_max", "score_mean"]].copy()
    distribution = distribution.rename(
        columns={
            "count": "final_count",
            "score_min": "final_score_min",
            "score_max": "final_score_max",
            "score_mean": "final_score_mean",
        }
    )
    jenks_counts = jenks_summary[["grade", "count", "score_min", "score_max"]].rename(
        columns={
            "count": "jenks_count",
            "score_min": "jenks_score_min",
            "score_max": "jenks_score_max",
        }
    )
    distribution = distribution.merge(jenks_counts, on="grade", how="left")
    distribution["grade"] = pd.Categorical(distribution["grade"], categories=GRADE_ORDER, ordered=True)
    distribution = distribution.sort_values("grade").reset_index(drop=True)
    distribution["grade"] = distribution["grade"].astype(str)

    summary = pd.DataFrame(
        [
            {"metric": "final_score_field", "value": "S_rank_v2"},
            {"metric": "final_grade_method", "value": "KMeans(k=5)"},
            {"metric": "robustness_check_method", "value": "Jenks natural breaks DP"},
            {"metric": "kmeans_silhouette_score", "value": silhouette},
            {"metric": "kmeans_jenks_consistency", "value": consistency},
        ]
    )
    return distribution, summary


def grade_interval_table(summary: pd.DataFrame, method: str) -> pd.DataFrame:
    """Build a readable score interval table for a grading method."""
    table = summary[["grade", "count", "score_min", "score_max", "score_mean"]].copy()
    table.insert(0, "method", method)
    table["score_interval"] = table.apply(
        lambda row: "" if pd.isna(row["score_min"]) else f"{row['score_min']:.6f} - {row['score_max']:.6f}",
        axis=1,
    )
    return table


def selected_output_columns(final_table: pd.DataFrame) -> list[str]:
    """Return final ranking output columns that exist in the table."""
    preferred = [
        "paper_id",
        "filename",
        "S_base",
        "S_BT",
        "S_BT_scaled",
        "S_rank_v2",
        "rank_base",
        "rank_final",
        "grade_final",
        "grade_kmeans",
        "grade_jenks",
        "kmeans_cluster",
        "kmeans_center",
    ]
    return [column for column in preferred if column in final_table.columns]


def log_summary(
    logger,
    final_table: pd.DataFrame,
    distribution: pd.DataFrame,
    silhouette: float,
    consistency: float,
    boundary_nearby: pd.DataFrame,
    comparison: pd.DataFrame,
) -> None:
    """Write the required result summary to the Step 8 log."""
    logger.info("Final score field used: S_rank_v2")
    logger.info(
        "Final S_rank_v2 range: %.6f - %.6f",
        float(final_table["S_rank_v2"].min()),
        float(final_table["S_rank_v2"].max()),
    )
    logger.info("Grade distribution:")
    for _, row in distribution.iterrows():
        logger.info(
            "  %s: %d papers, score interval %.6f - %.6f",
            row["grade"],
            int(row["final_count"]),
            float(row["final_score_min"]),
            float(row["final_score_max"]),
        )
    logger.info("KMeans silhouette score: %.6f", silhouette)
    logger.info("KMeans vs Jenks consistency: %.6f", consistency)

    top5 = final_table.sort_values("rank_final").head(5)
    bottom5 = final_table.sort_values("rank_final").tail(5)
    logger.info("Top 5 final ranking:")
    for _, row in top5.iterrows():
        logger.info(
            "  %d. %s %s S_rank_v2=%.6f grade=%s",
            int(row["rank_final"]),
            row["paper_id"],
            row["filename"],
            float(row["S_rank_v2"]),
            row["grade_final"],
        )
    logger.info("Bottom 5 final ranking:")
    for _, row in bottom5.iterrows():
        logger.info(
            "  %d. %s %s S_rank_v2=%.6f grade=%s",
            int(row["rank_final"]),
            row["paper_id"],
            row["filename"],
            float(row["S_rank_v2"]),
            row["grade_final"],
        )

    if boundary_nearby.empty:
        logger.info("No papers were found near KMeans grade boundaries.")
    else:
        logger.info("Boundary-nearby papers: %d", len(boundary_nearby))
        for _, row in boundary_nearby.head(20).iterrows():
            logger.info(
                "  %s rank=%d score=%.6f grade=%s boundary=%s distance=%.6f",
                row["paper_id"],
                int(row["rank_final"]),
                float(row["S_rank_v2"]),
                row["grade_kmeans"],
                row["boundary_between"],
                float(row["distance_to_boundary"]),
            )

    if consistency < 0.80:
        differences = comparison.loc[~comparison["is_same_grade"]]
        logger.warning(
            "KMeans and Jenks differ materially; inconsistent papers: %s",
            differences["paper_id"].tolist(),
        )
    else:
        differences = comparison.loc[~comparison["is_same_grade"]]
        logger.info("KMeans/Jenks inconsistent papers: %s", differences["paper_id"].tolist())

    logger.info("Writing suggestions for Problem 1:")
    logger.info("  1. 论文中明确说明最终分级使用 Step 7C 的 S_BT_scaled + lambda=0.85 融合分。")
    logger.info("  2. 说明 KMeans 为主分级方法，Jenks 自然断点作为稳健性验证。")
    logger.info("  3. 对等级边界附近论文在正文中保留解释，避免给出过度绝对化结论。")
    logger.info("  4. 可结合雷达图解释不同等级论文在结构、逻辑、建模、结果、写作五方面的差异。")


def main() -> None:
    """Run Step 8 grade classification and visualization."""
    config = load_config(PROJECT_ROOT / "config.yaml")
    paths = config.get("paths", {})

    tables_dir = resolve_project_path(paths.get("output_tables_dir", "output/tables"))
    charts_dir = resolve_project_path(paths.get("output_charts_dir", "output/charts"))
    logs_dir = resolve_project_path(paths.get("output_logs_dir", "output/logs"))
    logger = setup_grade_logger(logs_dir / "grade_visualization.log")

    logger.info("Starting Step 8 grade classification and visualization.")
    rank_fusion_v2_path = tables_dir / "appendix1_rank_fusion_v2.xlsx"
    feature_path = tables_dir / "appendix1_features_normalized.xlsx"
    weight_path = tables_dir / "appendix1_weights_ahp_entropy.xlsx"
    lambda_sensitivity_path = tables_dir / "bt_lambda_sensitivity.xlsx"
    rank_change_audit_path = tables_dir / "bt_rank_change_audit.xlsx"

    # Read required audit inputs to verify they exist and preserve traceability.
    pd.read_excel(lambda_sensitivity_path, sheet_name="recommendation")
    pd.read_excel(rank_change_audit_path, sheet_name="large_rank_changes")

    scores, score_column = read_final_scores(rank_fusion_v2_path)
    if score_column != "S_rank_v2":
        scores = scores.rename(columns={score_column: "S_rank_v2"})
        score_column = "S_rank_v2"

    kmeans_result, kmeans_summary, silhouette = classify_with_kmeans(scores, score_column)
    jenks_result, jenks_summary, jenks_breaks = classify_with_jenks(scores, score_column)
    final_table = merge_classifications(scores, kmeans_result, jenks_result)
    final_table = final_table[selected_output_columns(final_table)].copy()

    consistency, comparison = compare_grade_methods(final_table)
    boundary_nearby = find_boundary_nearby_papers(final_table, score_column)
    distribution, method_summary = make_distribution_table(
        kmeans_summary=kmeans_summary,
        jenks_summary=jenks_summary,
        silhouette=silhouette,
        consistency=consistency,
    )
    jenks_break_table = pd.DataFrame(
        {
            "break_index": list(range(len(jenks_breaks))),
            "break_value": jenks_breaks,
            "note": ["min"] + ["class upper bound"] * (len(jenks_breaks) - 2) + ["max"],
        }
    )

    feature_table = read_feature_table(feature_path)
    weight_table = read_weight_table(weight_path)
    chart_paths = {
        "final_ranking_bar": charts_dir / "final_ranking_bar.png",
        "grade_distribution": charts_dir / "grade_distribution.png",
        "high_mid_low_radar": charts_dir / "high_mid_low_radar.png",
        "final_score_histogram": charts_dir / "final_score_histogram.png",
        "kmeans_grade_scatter": charts_dir / "kmeans_grade_scatter.png",
        "kmeans_jenks_comparison": charts_dir / "kmeans_jenks_comparison.png",
    }
    visualization_outputs = create_problem1_visualizations(
        final_table=final_table,
        grade_summary=distribution.rename(
            columns={
                "final_count": "count",
                "final_score_min": "score_min",
                "final_score_max": "score_max",
                "final_score_mean": "score_mean",
            }
        ),
        feature_table=feature_table,
        weight_table=weight_table,
        chart_paths=chart_paths,
    )
    primary_scores = visualization_outputs["primary_scores"]

    final_ranking_path = write_workbook(
        tables_dir / "final_problem1_ranking.xlsx",
        {
            "final_ranking": final_table,
            "method_summary": method_summary,
            "grade_distribution": distribution,
            "boundary_nearby": boundary_nearby,
            "primary_scores": primary_scores,
        },
    )
    grade_distribution_path = write_workbook(
        tables_dir / "grade_distribution.xlsx",
        {
            "grade_distribution": distribution,
            "kmeans_intervals": grade_interval_table(kmeans_summary, "KMeans"),
            "jenks_intervals": grade_interval_table(jenks_summary, "Jenks"),
            "method_summary": method_summary,
        },
    )
    kmeans_details_path = write_workbook(
        tables_dir / "kmeans_grade_details.xlsx",
        {
            "kmeans_paper_grades": final_table[
                [
                    "paper_id",
                    "filename",
                    "S_rank_v2",
                    "rank_final",
                    "kmeans_cluster",
                    "kmeans_center",
                    "grade_kmeans",
                ]
            ],
            "kmeans_grade_summary": kmeans_summary,
            "boundary_nearby": boundary_nearby,
        },
    )
    jenks_details_path = write_workbook(
        tables_dir / "jenks_grade_details.xlsx",
        {
            "jenks_paper_grades": final_table[
                ["paper_id", "filename", "S_rank_v2", "rank_final", "grade_kmeans", "grade_jenks"]
            ],
            "jenks_grade_summary": jenks_summary,
            "jenks_breaks": jenks_break_table,
            "kmeans_jenks_comparison": comparison,
        },
    )

    log_summary(
        logger=logger,
        final_table=final_table,
        distribution=distribution,
        silhouette=silhouette,
        consistency=consistency,
        boundary_nearby=boundary_nearby,
        comparison=comparison,
    )
    logger.info("Saved final workbooks and charts.")
    logger.info("Finished Step 8. Problem 2 was not run.")

    print("Step 8 grade classification and visualization finished.")
    print(f"Final score field used: S_rank_v2")
    print(f"Final score range: {final_table['S_rank_v2'].min():.6f} - {final_table['S_rank_v2'].max():.6f}")
    print("Grade distribution:")
    for _, row in distribution.iterrows():
        print(f"  {row['grade']}: {int(row['final_count'])}")
    print(f"KMeans silhouette score: {silhouette:.6f}")
    print(f"KMeans-Jenks consistency: {consistency:.6f}")
    print("Top 5:")
    for _, row in final_table.head(5).iterrows():
        print(
            f"  {int(row['rank_final'])}. {row['paper_id']} {row['filename']} "
            f"{row['S_rank_v2']:.6f} {row['grade_final']}"
        )
    print("Bottom 5:")
    for _, row in final_table.tail(5).iterrows():
        print(
            f"  {int(row['rank_final'])}. {row['paper_id']} {row['filename']} "
            f"{row['S_rank_v2']:.6f} {row['grade_final']}"
        )
    print("Output tables:")
    for path in [final_ranking_path, grade_distribution_path, kmeans_details_path, jenks_details_path]:
        print(f"  {path}")
    print("Output charts:")
    for key, value in visualization_outputs.items():
        if key != "primary_scores":
            print(f"  {value}")
    print(f"Log: {logs_dir / 'grade_visualization.log'}")


if __name__ == "__main__":
    main()
