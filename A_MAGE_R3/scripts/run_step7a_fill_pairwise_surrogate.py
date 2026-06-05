"""Fill pairwise comparison template as a surrogate human reviewer.

This script only fills the winner and reason columns in a copied workbook:
output/tables/pairwise_comparison_filled.xlsx

It does not fit a Bradley-Terry model and does not use S_base directly to
choose winners. Pairwise decisions are based on five rubric dimensions derived
from the extracted feature table: model quality, result validation, logic,
symbol/formula standardization, and writing/application quality.
"""

from __future__ import annotations

from pathlib import Path
import re
import shutil

import numpy as np
import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def normalize_id(value) -> str:
    """Normalize paper IDs to two-digit strings."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    match = re.search(r"(\d+)", text)
    return match.group(1).zfill(2) if match else text


def feature_column(columns: list[str], prefix: str) -> str:
    """Find a feature column by its indicator prefix such as I09."""
    for column in columns:
        if isinstance(column, str) and column.startswith(prefix):
            return column
    raise KeyError(f"Feature column not found: {prefix}")


def weighted_sum(features: pd.DataFrame, specs: list[tuple[str, float]]) -> pd.Series:
    """Calculate weighted feature score for one review dimension."""
    score = pd.Series(0.0, index=features.index)
    for column, weight in specs:
        score = score + features[column] * weight
    return score


def build_review_scores(norm_table: pd.DataFrame) -> pd.DataFrame:
    """Build five-dimensional surrogate human review scores."""
    feature_cols = [column for column in norm_table.columns if isinstance(column, str) and column.startswith("I")]
    col = {f"I{index:02d}": feature_column(feature_cols, f"I{index:02d}_") for index in range(1, 22)}

    norm_table = norm_table.copy()
    norm_table["paper_id_norm"] = norm_table["paper_id"].apply(normalize_id)
    features = norm_table.set_index("paper_id_norm")[feature_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)

    review = pd.DataFrame(index=features.index)
    review["model"] = weighted_sum(
        features,
        [
            (col["I09"], 0.18),
            (col["I10"], 0.18),
            (col["I11"], 0.20),
            (col["I12"], 0.24),
            (col["I13"], 0.20),
        ],
    )
    review["result"] = weighted_sum(
        features,
        [
            (col["I14"], 0.35),
            (col["I15"], 0.25),
            (col["I16"], 0.20),
            (col["I17"], 0.20),
        ],
    )
    review["logic"] = weighted_sum(
        features,
        [
            (col["I05"], 0.30),
            (col["I06"], 0.28),
            (col["I07"], 0.20),
            (col["I08"], 0.22),
        ],
    )
    review["symbol"] = weighted_sum(
        features,
        [
            (col["I10"], 0.25),
            (col["I11"], 0.35),
            (col["I12"], 0.30),
            (col["I03"], 0.10),
        ],
    )
    review["writing"] = weighted_sum(
        features,
        [
            (col["I01"], 0.18),
            (col["I02"], 0.14),
            (col["I03"], 0.13),
            (col["I18"], 0.14),
            (col["I19"], 0.14),
            (col["I20"], 0.13),
            (col["I21"], 0.14),
        ],
    )

    review["review_score"] = (
        0.35 * review["model"]
        + 0.25 * review["result"]
        + 0.20 * review["logic"]
        + 0.12 * review["symbol"]
        + 0.08 * review["writing"]
    )
    return review


RUBRIC_LABELS = {
    "model": "模型建立更合理",
    "result": "结果分析和验证更充分",
    "logic": "问题理解与逻辑链更清楚",
    "symbol": "变量公式和符号说明更规范",
    "writing": "写作结构和应用表达更完整",
}

PAIR_REASONS = {
    ("model", "result"): "模型建立更合理，结果验证更充分",
    ("model", "logic"): "模型建立更合理，逻辑链更清楚",
    ("model", "symbol"): "模型构建更充分，变量公式更规范",
    ("model", "writing"): "模型质量更高，写作结构更完整",
    ("result", "logic"): "结果分析更充分，逻辑表达更清楚",
    ("result", "symbol"): "结果验证更充分，图表和公式更规范",
    ("result", "writing"): "结果分析更完整，写作表达更规范",
    ("logic", "symbol"): "逻辑链更清楚，变量符号更规范",
    ("logic", "writing"): "问题理解更到位，写作结构更清晰",
    ("symbol", "writing"): "变量公式更规范，图表写作更清楚",
}


def winner_reason(winner_pid: str, loser_pid: str, review: pd.DataFrame) -> str:
    """Generate a concise reason for the selected winner."""
    diffs = {
        dim: float(review.loc[winner_pid, dim] - review.loc[loser_pid, dim])
        for dim in ["model", "result", "logic", "symbol", "writing"]
    }
    positive = [dim for dim, value in sorted(diffs.items(), key=lambda item: item[1], reverse=True) if value > 0.015]
    if len(positive) >= 2:
        key = tuple(positive[:2])
        return PAIR_REASONS.get(key, PAIR_REASONS.get(tuple(reversed(key)), f"{RUBRIC_LABELS[positive[0]]}，{RUBRIC_LABELS[positive[1]]}"))
    if len(positive) == 1:
        return f"{RUBRIC_LABELS[positive[0]]}，综合表现更稳定"
    return "综合结构和模型表达略优"


def tie_reason(pid_i: str, pid_j: str, review: pd.DataFrame) -> str:
    """Generate a concise reason for a tie."""
    readable = {
        "model": "模型质量",
        "result": "结果验证",
        "logic": "逻辑表达",
        "symbol": "变量公式",
        "writing": "写作结构",
    }
    diffs = {
        dim: abs(float(review.loc[pid_i, dim] - review.loc[pid_j, dim]))
        for dim in ["model", "result", "logic", "symbol", "writing"]
    }
    closest = sorted(diffs, key=diffs.get)[:2]
    return f"两文{readable[closest[0]]}和{readable[closest[1]]}接近，整体难以区分"


def decide_pair(pid_i: str, pid_j: str, review: pd.DataFrame) -> tuple[str, str]:
    """Decide a pairwise winner without using TOPSIS S_base."""
    score_i = float(review.loc[pid_i, "review_score"])
    score_j = float(review.loc[pid_j, "review_score"])
    diff = score_i - score_j
    dim_diff = {
        dim: float(review.loc[pid_i, dim] - review.loc[pid_j, dim])
        for dim in ["model", "result", "logic", "symbol", "writing"]
    }
    i_leads = sum(1 for value in dim_diff.values() if value > 0.025)
    j_leads = sum(1 for value in dim_diff.values() if value < -0.025)

    if abs(diff) < 0.018:
        return "tie", tie_reason(pid_i, pid_j, review)
    if abs(diff) < 0.040 and abs(i_leads - j_leads) <= 1:
        return "tie", tie_reason(pid_i, pid_j, review)
    if diff > 0:
        return "i", winner_reason(pid_i, pid_j, review)
    return "j", winner_reason(pid_j, pid_i, review)


def main() -> None:
    """Fill the pairwise comparison template copy."""
    template_path = PROJECT_ROOT / "output/tables/pairwise_comparison_template.xlsx"
    filled_path = PROJECT_ROOT / "output/tables/pairwise_comparison_filled.xlsx"
    norm_path = PROJECT_ROOT / "output/tables/appendix1_features_normalized.xlsx"

    shutil.copyfile(template_path, filled_path)
    norm = pd.read_excel(norm_path)
    review = build_review_scores(norm)

    workbook = load_workbook(filled_path)
    sheet = workbook["pairwise_template"]
    headers = {cell.value: index for index, cell in enumerate(sheet[1], start=1)}
    winner_col = headers["winner"]
    reason_col = headers["reason"]
    paper_i_col = headers["paper_i"]
    paper_j_col = headers["paper_j"]

    for row_index in range(2, sheet.max_row + 1):
        pid_i = normalize_id(sheet.cell(row_index, paper_i_col).value)
        pid_j = normalize_id(sheet.cell(row_index, paper_j_col).value)
        winner, reason = decide_pair(pid_i, pid_j, review)
        sheet.cell(row_index, winner_col).value = winner
        sheet.cell(row_index, reason_col).value = reason

    workbook.save(filled_path)

    filled = pd.read_excel(filled_path, sheet_name="pairwise_template")
    valid = filled["winner"].isin(["i", "j", "tie"])
    counts = {paper_id: 0 for paper_id in review.index}
    for _, row in filled[valid].iterrows():
        counts[normalize_id(row["paper_i"])] += 1
        counts[normalize_id(row["paper_j"])] += 1
    undercovered = [paper_id for paper_id, count in counts.items() if count < 2]

    print("Filled pairwise comparison template saved.")
    print(f"Output: {filled_path}")
    print(f"Total pairs: {len(filled)}")
    print(f"Valid filled pairs: {int(valid.sum())}")
    print(f"Winner counts: {filled['winner'].value_counts(dropna=False).to_dict()}")
    print(f"Average appearances per paper: {float(np.mean(list(counts.values()))):.4f}")
    print(f"Minimum appearances per paper: {min(counts.values())}")
    print(f"Undercovered papers: {', '.join(undercovered) if undercovered else 'none'}")


if __name__ == "__main__":
    main()
