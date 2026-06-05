"""Step 8B: final consistency audit for Problem 1 outputs.

This script is intentionally read-only for Step 8 result files. It does not
retrain, rescore, regroup, or modify any original result workbook.
"""

from __future__ import annotations

from pathlib import Path
import logging
import re
import sys
from typing import Any

import numpy as np
import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

TABLES_DIR = PROJECT_ROOT / "output/tables"
CHARTS_DIR = PROJECT_ROOT / "output/charts"
LOGS_DIR = PROJECT_ROOT / "output/logs"
PAPER_SECTIONS_DIR = PROJECT_ROOT / "paper_sections"

REQUIRED_CHARTS = [
    "final_ranking_bar.png",
    "grade_distribution.png",
    "high_mid_low_radar.png",
    "final_score_histogram.png",
    "kmeans_grade_scatter.png",
    "kmeans_jenks_comparison.png",
]

EXPECTED_GRADE_ORDER = ["优秀", "良好", "中等", "及格", "不及格"]
SCORE_TOLERANCE = 1e-6


def setup_logger(log_path: Path) -> logging.Logger:
    """Create an audit logger."""
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("A_MAGE_R3.problem1_final_audit")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    handler = logging.FileHandler(log_path, mode="w", encoding="utf-8-sig")
    handler.setFormatter(
        logging.Formatter(
            fmt="%(asctime)s [%(levelname)s] %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    logger.addHandler(handler)
    return logger


def normalize_id(value: Any) -> str:
    """Normalize paper IDs to two-digit strings."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    match = re.search(r"(\d+)", text)
    return match.group(1).zfill(2) if match else text.zfill(2)


def pass_fail(condition: bool) -> str:
    """Return PASS or FAIL."""
    return "PASS" if bool(condition) else "FAIL"


def read_required_inputs() -> dict[str, pd.DataFrame]:
    """Read all required input workbooks and sheets."""
    inputs = {
        "final_ranking": pd.read_excel(TABLES_DIR / "final_problem1_ranking.xlsx", sheet_name="final_ranking"),
        "final_method_summary": pd.read_excel(
            TABLES_DIR / "final_problem1_ranking.xlsx", sheet_name="method_summary"
        ),
        "grade_distribution": pd.read_excel(
            TABLES_DIR / "grade_distribution.xlsx", sheet_name="grade_distribution"
        ),
        "grade_method_summary": pd.read_excel(
            TABLES_DIR / "grade_distribution.xlsx", sheet_name="method_summary"
        ),
        "kmeans_paper_grades": pd.read_excel(
            TABLES_DIR / "kmeans_grade_details.xlsx", sheet_name="kmeans_paper_grades"
        ),
        "kmeans_grade_summary": pd.read_excel(
            TABLES_DIR / "kmeans_grade_details.xlsx", sheet_name="kmeans_grade_summary"
        ),
        "jenks_paper_grades": pd.read_excel(
            TABLES_DIR / "jenks_grade_details.xlsx", sheet_name="jenks_paper_grades"
        ),
        "jenks_grade_summary": pd.read_excel(
            TABLES_DIR / "jenks_grade_details.xlsx", sheet_name="jenks_grade_summary"
        ),
        "kmeans_jenks_comparison": pd.read_excel(
            TABLES_DIR / "jenks_grade_details.xlsx", sheet_name="kmeans_jenks_comparison"
        ),
        "rank_fusion_v2": pd.read_excel(
            TABLES_DIR / "appendix1_rank_fusion_v2.xlsx", sheet_name="rank_fusion_v2"
        ),
        "topsis_scores": pd.read_excel(TABLES_DIR / "appendix1_topsis_scores.xlsx", sheet_name="topsis_scores"),
        "combined_weights": pd.read_excel(
            TABLES_DIR / "appendix1_weights_ahp_entropy.xlsx", sheet_name="combined_weights"
        ),
    }

    for key in [
        "final_ranking",
        "kmeans_paper_grades",
        "jenks_paper_grades",
        "kmeans_jenks_comparison",
        "rank_fusion_v2",
        "topsis_scores",
    ]:
        if "paper_id" in inputs[key].columns:
            inputs[key] = inputs[key].copy()
            inputs[key]["paper_id"] = inputs[key]["paper_id"].apply(normalize_id)

    return inputs


def make_file_status() -> pd.DataFrame:
    """Check required input and output-support files exist."""
    files = [
        TABLES_DIR / "final_problem1_ranking.xlsx",
        TABLES_DIR / "grade_distribution.xlsx",
        TABLES_DIR / "kmeans_grade_details.xlsx",
        TABLES_DIR / "jenks_grade_details.xlsx",
        TABLES_DIR / "appendix1_rank_fusion_v2.xlsx",
        TABLES_DIR / "appendix1_topsis_scores.xlsx",
        TABLES_DIR / "appendix1_weights_ahp_entropy.xlsx",
        LOGS_DIR / "grade_visualization.log",
    ] + [CHARTS_DIR / chart_name for chart_name in REQUIRED_CHARTS]

    rows = []
    for file_path in files:
        rows.append(
            {
                "file": str(file_path.relative_to(PROJECT_ROOT)),
                "exists": file_path.exists(),
                "size_bytes": file_path.stat().st_size if file_path.exists() else 0,
                "status": pass_fail(file_path.exists() and file_path.stat().st_size > 0)
                if file_path.exists()
                else "FAIL",
            }
        )
    return pd.DataFrame(rows)


def audit_row_count_and_ids(
    final_ranking: pd.DataFrame,
    rank_fusion_v2: pd.DataFrame,
    topsis_scores: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Check row count, duplicates, and missing IDs."""
    final_ids = final_ranking["paper_id"].astype(str)
    fusion_ids = set(rank_fusion_v2["paper_id"].astype(str))
    topsis_ids = set(topsis_scores["paper_id"].astype(str))
    expected_ids = set(f"{index:02d}" for index in range(1, 31))
    reference_ids = expected_ids | fusion_ids | topsis_ids

    duplicate_ids = sorted(final_ids[final_ids.duplicated()].unique().tolist())
    missing_from_expected = sorted(reference_ids - set(final_ids))
    extra_in_final = sorted(set(final_ids) - reference_ids)

    checks = [
        {
            "check_id": "row_count_30",
            "item": "final_problem1_ranking.xlsx contains 30 papers",
            "status": pass_fail(len(final_ranking) == 30),
            "details": f"actual_rows={len(final_ranking)}",
        },
        {
            "check_id": "paper_id_no_duplicate",
            "item": "paper_id has no duplicates",
            "status": pass_fail(len(duplicate_ids) == 0),
            "details": f"duplicate_ids={duplicate_ids}",
        },
        {
            "check_id": "paper_id_no_missing",
            "item": "paper_id has no missing IDs against 01-30/fusion/TOPSIS references",
            "status": pass_fail(len(missing_from_expected) == 0 and len(extra_in_final) == 0),
            "details": f"missing={missing_from_expected}; extra={extra_in_final}",
        },
    ]
    id_detail = pd.DataFrame(
        {
            "paper_id": sorted(reference_ids),
            "in_final": [paper_id in set(final_ids) for paper_id in sorted(reference_ids)],
            "in_rank_fusion_v2": [paper_id in fusion_ids for paper_id in sorted(reference_ids)],
            "in_topsis_scores": [paper_id in topsis_ids for paper_id in sorted(reference_ids)],
            "is_expected_01_30": [paper_id in expected_ids for paper_id in sorted(reference_ids)],
        }
    )
    return pd.DataFrame(checks), id_detail


def audit_score_consistency(final_ranking: pd.DataFrame, rank_fusion_v2: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Check S_rank_v2 consistency between final ranking and Step 7C fusion table."""
    merged = final_ranking[["paper_id", "filename", "S_rank_v2"]].merge(
        rank_fusion_v2[["paper_id", "S_rank_v2", "S_base", "S_BT_scaled", "rank_fused_v2"]],
        on="paper_id",
        suffixes=("_final", "_fusion_v2"),
        how="outer",
        indicator=True,
    )
    for column in ["S_rank_v2_final", "S_rank_v2_fusion_v2"]:
        merged[column] = pd.to_numeric(merged[column], errors="coerce")
    merged["abs_diff"] = (merged["S_rank_v2_final"] - merged["S_rank_v2_fusion_v2"]).abs()
    merged["is_consistent"] = (merged["_merge"] == "both") & (merged["abs_diff"] <= SCORE_TOLERANCE)
    inconsistent = merged.loc[~merged["is_consistent"]]
    check = pd.DataFrame(
        [
            {
                "check_id": "s_rank_v2_consistency",
                "item": "S_rank_v2 matches appendix1_rank_fusion_v2.xlsx",
                "status": pass_fail(inconsistent.empty),
                "details": f"inconsistent_paper_ids={inconsistent['paper_id'].tolist()}",
            }
        ]
    )
    return check, merged.sort_values("paper_id").reset_index(drop=True)


def audit_rank_order(final_ranking: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Check rank_final follows descending S_rank_v2."""
    table = final_ranking.copy()
    table["S_rank_v2"] = pd.to_numeric(table["S_rank_v2"], errors="coerce")
    table["rank_final"] = pd.to_numeric(table["rank_final"], errors="coerce")
    expected = table.sort_values(["S_rank_v2", "paper_id"], ascending=[False, True]).reset_index(drop=True)
    expected["expected_rank_final"] = np.arange(1, len(expected) + 1)
    rank_detail = table.merge(
        expected[["paper_id", "expected_rank_final"]],
        on="paper_id",
        how="left",
    )
    rank_detail["rank_diff"] = rank_detail["rank_final"] - rank_detail["expected_rank_final"]
    bad = rank_detail.loc[rank_detail["rank_diff"] != 0]
    rank_set_ok = set(rank_detail["rank_final"].dropna().astype(int)) == set(range(1, len(rank_detail) + 1))
    check = pd.DataFrame(
        [
            {
                "check_id": "rank_final_descending",
                "item": "rank_final follows S_rank_v2 descending order",
                "status": pass_fail(bad.empty and rank_set_ok),
                "details": f"rank_mismatch_ids={bad['paper_id'].tolist()}; rank_set_ok={rank_set_ok}",
            }
        ]
    )
    return check, rank_detail.sort_values("rank_final").reset_index(drop=True)


def audit_grade_distribution(grade_distribution: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Check grade distribution count sum and grade labels."""
    distribution = grade_distribution.copy()
    distribution["final_count"] = pd.to_numeric(distribution["final_count"], errors="coerce")
    count_sum = int(distribution["final_count"].sum())
    grades = distribution["grade"].astype(str).tolist()
    check = pd.DataFrame(
        [
            {
                "check_id": "grade_distribution_sum",
                "item": "grade_distribution final counts sum to 30",
                "status": pass_fail(count_sum == 30),
                "details": f"count_sum={count_sum}",
            },
            {
                "check_id": "grade_distribution_labels",
                "item": "grade_distribution has five expected grade labels",
                "status": pass_fail(grades == EXPECTED_GRADE_ORDER),
                "details": f"grades={grades}",
            },
        ]
    )
    return check, distribution


def audit_grade_consistency(
    final_ranking: pd.DataFrame,
    kmeans_paper_grades: pd.DataFrame,
    jenks_paper_grades: pd.DataFrame,
    kmeans_jenks_comparison: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Check grade_final, grade_kmeans, and grade_jenks consistency."""
    detail = final_ranking[
        ["paper_id", "filename", "S_rank_v2", "rank_final", "grade_final", "grade_kmeans", "grade_jenks"]
    ].copy()
    detail["final_equals_kmeans"] = detail["grade_final"] == detail["grade_kmeans"]
    detail["kmeans_equals_jenks"] = detail["grade_kmeans"] == detail["grade_jenks"]
    detail["all_three_same"] = detail["final_equals_kmeans"] & detail["kmeans_equals_jenks"]

    kmeans_join = detail[["paper_id", "grade_kmeans"]].merge(
        kmeans_paper_grades[["paper_id", "grade_kmeans"]].rename(columns={"grade_kmeans": "grade_kmeans_detail"}),
        on="paper_id",
        how="left",
    )
    jenks_join = detail[["paper_id", "grade_jenks"]].merge(
        jenks_paper_grades[["paper_id", "grade_jenks"]].rename(columns={"grade_jenks": "grade_jenks_detail"}),
        on="paper_id",
        how="left",
    )
    kmeans_detail_ok = bool((kmeans_join["grade_kmeans"] == kmeans_join["grade_kmeans_detail"]).all())
    jenks_detail_ok = bool((jenks_join["grade_jenks"] == jenks_join["grade_jenks_detail"]).all())
    consistency = float(detail["kmeans_equals_jenks"].mean()) if len(detail) else np.nan

    comparison_consistency = np.nan
    if "is_same_grade" in kmeans_jenks_comparison.columns:
        comparison_consistency = float(kmeans_jenks_comparison["is_same_grade"].astype(bool).mean())

    inconsistent_ids = detail.loc[~detail["all_three_same"], "paper_id"].tolist()
    checks = pd.DataFrame(
        [
            {
                "check_id": "final_kmeans_jenks_grade_consistency",
                "item": "grade_final, grade_kmeans, and grade_jenks are identical in final ranking",
                "status": pass_fail(len(inconsistent_ids) == 0),
                "details": f"inconsistent_ids={inconsistent_ids}",
            },
            {
                "check_id": "detail_workbook_grade_consistency",
                "item": "KMeans/Jenks detail workbooks match final ranking",
                "status": pass_fail(kmeans_detail_ok and jenks_detail_ok),
                "details": f"kmeans_detail_ok={kmeans_detail_ok}; jenks_detail_ok={jenks_detail_ok}",
            },
            {
                "check_id": "kmeans_jenks_consistency_rate",
                "item": "KMeans and Jenks consistency rate equals 1.0",
                "status": pass_fail(abs(consistency - 1.0) <= SCORE_TOLERANCE),
                "details": f"computed_consistency={consistency:.6f}; detail_sheet_consistency={comparison_consistency:.6f}",
            },
        ]
    )
    return checks, detail.sort_values("rank_final").reset_index(drop=True)


def parse_top_bottom_from_log(log_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse top 5 and bottom 5 result lines from Step 8 log."""
    if not log_path.exists():
        return pd.DataFrame(), pd.DataFrame()
    lines = log_path.read_text(encoding="utf-8-sig", errors="replace").splitlines()
    mode = None
    rows: dict[str, list[dict[str, Any]]] = {"top": [], "bottom": []}
    pattern = re.compile(
        r"\s(\d+)\.\s+(\d+)\s+(\S+)\s+S_rank_v2=([0-9.]+)\s+grade=(\S+)"
    )
    for line in lines:
        if "Top 5 final ranking:" in line:
            mode = "top"
            continue
        if "Bottom 5 final ranking:" in line:
            mode = "bottom"
            continue
        if mode not in rows:
            continue
        match = pattern.search(line)
        if match:
            rows[mode].append(
                {
                    "rank_final": int(match.group(1)),
                    "paper_id": normalize_id(match.group(2)),
                    "filename": match.group(3),
                    "S_rank_v2_log": float(match.group(4)),
                    "grade_log": match.group(5),
                }
            )
            if len(rows[mode]) >= 5:
                mode = None
    return pd.DataFrame(rows["top"]), pd.DataFrame(rows["bottom"])


def audit_top_bottom_against_log(final_ranking: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Check final top 5 and bottom 5 against grade_visualization.log."""
    top_log, bottom_log = parse_top_bottom_from_log(LOGS_DIR / "grade_visualization.log")
    final_top = final_ranking.sort_values("rank_final").head(5)[
        ["rank_final", "paper_id", "filename", "S_rank_v2", "grade_final"]
    ]
    final_bottom = final_ranking.sort_values("rank_final").tail(5)[
        ["rank_final", "paper_id", "filename", "S_rank_v2", "grade_final"]
    ]

    def compare(final_part: pd.DataFrame, log_part: pd.DataFrame, label: str) -> tuple[bool, pd.DataFrame]:
        if log_part.empty:
            detail = final_part.copy()
            detail["part"] = label
            detail["log_found"] = False
            return False, detail
        detail = final_part.merge(log_part, on=["rank_final", "paper_id", "filename"], how="left")
        detail["part"] = label
        detail["log_found"] = detail["S_rank_v2_log"].notna()
        detail["score_matches_log"] = (detail["S_rank_v2"] - detail["S_rank_v2_log"]).abs() <= 1e-5
        detail["grade_matches_log"] = detail["grade_final"] == detail["grade_log"]
        ok = bool((detail["log_found"] & detail["score_matches_log"] & detail["grade_matches_log"]).all())
        return ok, detail

    top_ok, top_detail = compare(final_top, top_log, "top5")
    bottom_ok, bottom_detail = compare(final_bottom, bottom_log, "bottom5")
    detail = pd.concat([top_detail, bottom_detail], ignore_index=True)
    check = pd.DataFrame(
        [
            {
                "check_id": "top_bottom_log_consistency",
                "item": "Final top 5 and bottom 5 match grade_visualization.log",
                "status": pass_fail(top_ok and bottom_ok),
                "details": f"top_ok={top_ok}; bottom_ok={bottom_ok}",
            }
        ]
    )
    return check, detail


def audit_charts(file_status: pd.DataFrame) -> pd.DataFrame:
    """Check required chart files exist and are non-empty."""
    chart_status = file_status.loc[file_status["file"].str.startswith("output/charts/")].copy()
    missing = chart_status.loc[chart_status["status"] != "PASS", "file"].tolist()
    return pd.DataFrame(
        [
            {
                "check_id": "chart_files_exist",
                "item": "All final chart files exist and are non-empty",
                "status": pass_fail(len(missing) == 0),
                "details": f"missing_or_empty={missing}",
            }
        ]
    )


def audit_weights(combined_weights: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Lightly verify weight table structure used by final visualization inputs."""
    detail = combined_weights.copy()
    weight_sum = pd.to_numeric(detail.get("combined_weight"), errors="coerce").sum()
    indicator_count = int(detail["indicator"].nunique()) if "indicator" in detail.columns else 0
    required_columns = {"criterion", "indicator", "combined_weight"}
    missing_columns = sorted(required_columns - set(detail.columns))
    check = pd.DataFrame(
        [
            {
                "check_id": "weight_file_basic_validity",
                "item": "Weight table has 21 indicators and normalized combined weights",
                "status": pass_fail(indicator_count == 21 and abs(weight_sum - 1.0) <= 1e-6 and not missing_columns),
                "details": (
                    f"indicator_count={indicator_count}; combined_weight_sum={weight_sum:.12f}; "
                    f"missing_columns={missing_columns}"
                ),
            }
        ]
    )
    return check, detail


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> Path:
    """Write audit workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    if load_workbook is not None:
        workbook = load_workbook(path)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                letter = column[0].column_letter
                width = min(max(max(len(str(cell.value or "")) for cell in column[:200]) + 2, 10), 48)
                sheet.column_dimensions[letter].width = width
        workbook.save(path)
    return path


def write_markdown_report(
    path: Path,
    all_checks: pd.DataFrame,
    final_ranking: pd.DataFrame,
    grade_distribution: pd.DataFrame,
    consistency: float,
) -> Path:
    """Write a concise Markdown audit report for paper writing."""
    path.parent.mkdir(parents=True, exist_ok=True)
    failed = all_checks.loc[all_checks["status"] != "PASS"]
    conclusion = (
        "问题1结果文件一致，可进入论文写作阶段"
        if failed.empty
        else "问题1结果文件存在不一致，需先修正后再进入论文写作阶段"
    )

    top5 = final_ranking.sort_values("rank_final").head(5)
    bottom5 = final_ranking.sort_values("rank_final").tail(5)

    lines = [
        "# 问题1最终结果一致性审计报告",
        "",
        f"审计结论：**{conclusion}**",
        "",
        "## 审计范围",
        "",
        "- final_problem1_ranking.xlsx",
        "- grade_distribution.xlsx",
        "- kmeans_grade_details.xlsx",
        "- jenks_grade_details.xlsx",
        "- appendix1_rank_fusion_v2.xlsx",
        "- appendix1_topsis_scores.xlsx",
        "- appendix1_weights_ahp_entropy.xlsx",
        "",
        "## 关键结果",
        "",
        f"- 论文数量：{len(final_ranking)}",
        f"- S_rank_v2 范围：{final_ranking['S_rank_v2'].min():.6f} - {final_ranking['S_rank_v2'].max():.6f}",
        f"- KMeans 与 Jenks 一致率：{consistency:.6f}",
        "",
        "## 五级分布",
        "",
        "| 等级 | 数量 | 分数区间 |",
        "|---|---:|---|",
    ]
    for _, row in grade_distribution.iterrows():
        lines.append(
            f"| {row['grade']} | {int(row['final_count'])} | "
            f"{float(row['final_score_min']):.6f} - {float(row['final_score_max']):.6f} |"
        )

    lines += [
        "",
        "## 最终排名前5",
        "",
        "| 排名 | 论文 | S_rank_v2 | 等级 |",
        "|---:|---|---:|---|",
    ]
    for _, row in top5.iterrows():
        lines.append(
            f"| {int(row['rank_final'])} | {row['filename']} | "
            f"{float(row['S_rank_v2']):.6f} | {row['grade_final']} |"
        )

    lines += [
        "",
        "## 最终排名后5",
        "",
        "| 排名 | 论文 | S_rank_v2 | 等级 |",
        "|---:|---|---:|---|",
    ]
    for _, row in bottom5.iterrows():
        lines.append(
            f"| {int(row['rank_final'])} | {row['filename']} | "
            f"{float(row['S_rank_v2']):.6f} | {row['grade_final']} |"
        )

    lines += [
        "",
        "## 审计项",
        "",
        "| 检查项 | 状态 | 说明 |",
        "|---|---|---|",
    ]
    for _, row in all_checks.iterrows():
        lines.append(f"| {row['item']} | {row['status']} | {row['details']} |")

    if not failed.empty:
        lines += [
            "",
            "## 待修正问题",
            "",
        ]
        for _, row in failed.iterrows():
            lines.append(f"- {row['check_id']}：{row['details']}")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def main() -> None:
    """Run the final Problem 1 audit."""
    logger = setup_logger(LOGS_DIR / "problem1_final_audit.log")
    logger.info("Starting Step 8B final consistency audit. No models will be recalculated.")

    inputs = read_required_inputs()
    file_status = make_file_status()
    row_checks, id_detail = audit_row_count_and_ids(
        inputs["final_ranking"],
        inputs["rank_fusion_v2"],
        inputs["topsis_scores"],
    )
    score_check, score_detail = audit_score_consistency(inputs["final_ranking"], inputs["rank_fusion_v2"])
    rank_check, rank_detail = audit_rank_order(inputs["final_ranking"])
    distribution_check, distribution_detail = audit_grade_distribution(inputs["grade_distribution"])
    grade_checks, grade_detail = audit_grade_consistency(
        inputs["final_ranking"],
        inputs["kmeans_paper_grades"],
        inputs["jenks_paper_grades"],
        inputs["kmeans_jenks_comparison"],
    )
    top_bottom_check, top_bottom_detail = audit_top_bottom_against_log(inputs["final_ranking"])
    chart_check = audit_charts(file_status)
    weight_check, weight_detail = audit_weights(inputs["combined_weights"])

    all_checks = pd.concat(
        [
            row_checks,
            score_check,
            rank_check,
            distribution_check,
            grade_checks,
            top_bottom_check,
            chart_check,
            weight_check,
        ],
        ignore_index=True,
    )
    all_pass = bool((all_checks["status"] == "PASS").all())
    conclusion = (
        "问题1结果文件一致，可进入论文写作阶段"
        if all_pass
        else "问题1结果文件存在不一致，需先修正后再进入论文写作阶段"
    )
    summary = pd.DataFrame(
        [
            {"metric": "audit_status", "value": "PASS" if all_pass else "FAIL"},
            {"metric": "conclusion", "value": conclusion},
            {"metric": "paper_count", "value": len(inputs["final_ranking"])},
            {
                "metric": "S_rank_v2_min",
                "value": float(pd.to_numeric(inputs["final_ranking"]["S_rank_v2"], errors="coerce").min()),
            },
            {
                "metric": "S_rank_v2_max",
                "value": float(pd.to_numeric(inputs["final_ranking"]["S_rank_v2"], errors="coerce").max()),
            },
            {
                "metric": "kmeans_jenks_consistency",
                "value": float(grade_detail["kmeans_equals_jenks"].mean()),
            },
            {"metric": "failed_check_count", "value": int((all_checks["status"] != "PASS").sum())},
        ]
    )

    audit_path = write_workbook(
        TABLES_DIR / "problem1_final_audit.xlsx",
        {
            "summary": summary,
            "checks": all_checks,
            "file_status": file_status,
            "paper_id_integrity": id_detail,
            "score_consistency": score_detail,
            "rank_order_check": rank_detail,
            "grade_consistency": grade_detail,
            "distribution_check": distribution_detail,
            "top_bottom_log_check": top_bottom_detail,
            "weight_file_check": weight_detail,
        },
    )
    md_path = write_markdown_report(
        PAPER_SECTIONS_DIR / "problem1_result_summary.md",
        all_checks=all_checks,
        final_ranking=inputs["final_ranking"],
        grade_distribution=inputs["grade_distribution"],
        consistency=float(grade_detail["kmeans_equals_jenks"].mean()),
    )

    for _, row in all_checks.iterrows():
        log_method = logger.info if row["status"] == "PASS" else logger.error
        log_method("%s | %s | %s", row["status"], row["check_id"], row["details"])
    logger.info("Audit conclusion: %s", conclusion)
    logger.info("Audit workbook saved: %s", audit_path)
    logger.info("Markdown summary saved: %s", md_path)
    logger.info("Finished Step 8B. Original result files were not modified.")

    print("Step 8B final audit finished.")
    print(f"Audit status: {'PASS' if all_pass else 'FAIL'}")
    print(conclusion)
    print(f"Audit workbook: {audit_path}")
    print(f"Audit log: {LOGS_DIR / 'problem1_final_audit.log'}")
    print(f"Markdown summary: {md_path}")
    print("Failed checks:" if not all_pass else "Failed checks: none")
    if not all_pass:
        for _, row in all_checks.loc[all_checks["status"] != "PASS"].iterrows():
            print(f"  {row['check_id']}: {row['details']}")


if __name__ == "__main__":
    main()

