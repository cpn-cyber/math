"""Step 11: judge-driven revision audit for Problem 1.

This step responds to the external judge review. It does not change sealed
scores, ranks, grades, or Step 1-8B result files. It adds credibility checks,
external-anchor templates, and paper-writing evidence.
"""

from __future__ import annotations

from itertools import combinations
from pathlib import Path
import logging
import math
import re
import sys
from typing import Any

import numpy as np
import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:  # pragma: no cover
    load_workbook = None
    Alignment = Font = PatternFill = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

TABLES_DIR = PROJECT_ROOT / "output/tables"
LOGS_DIR = PROJECT_ROOT / "output/logs"
PAPER_SECTIONS_DIR = PROJECT_ROOT / "paper_sections"

LAMBDA_FINAL = 0.85
ALPHA_GRID = [0.4, 0.5, 0.6, 0.7]
GRADE_ORDER = ["优秀", "良好", "中等", "及格", "不及格"]


def setup_logger(path: Path) -> logging.Logger:
    """Create the Step 11 logger."""
    path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("A_MAGE_R3.problem1_judge_revision")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    handler = logging.FileHandler(path, mode="w", encoding="utf-8-sig")
    handler.setFormatter(
        logging.Formatter(
            fmt="%(asctime)s [%(levelname)s] %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    logger.addHandler(handler)
    return logger


def normalize_id(value: Any) -> str:
    """Normalize paper IDs as two-digit strings."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    match = re.search(r"(\d+)", text)
    return match.group(1).zfill(2) if match else text.zfill(2)


def indicator_number(column: str) -> int:
    """Return the numeric suffix of an indicator column."""
    match = re.match(r"I(\d+)", str(column))
    return int(match.group(1)) if match else 999


def rank_desc(scores: pd.Series) -> pd.Series:
    """Rank scores in descending order."""
    return scores.rank(ascending=False, method="first").astype(int)


def spearman_from_ranks(left: pd.Series, right: pd.Series) -> float:
    """Calculate Spearman correlation between two aligned rank columns."""
    left = pd.to_numeric(left, errors="coerce")
    right = pd.to_numeric(right, errors="coerce")
    valid = ~(left.isna() | right.isna())
    if valid.sum() < 2:
        return float("nan")
    return float(np.corrcoef(left[valid], right[valid])[0, 1])


def kendall_tau_from_ranks(left: pd.Series, right: pd.Series) -> float:
    """Calculate Kendall tau-a for complete rankings."""
    left = pd.to_numeric(left, errors="coerce").to_numpy()
    right = pd.to_numeric(right, errors="coerce").to_numpy()
    valid = ~(np.isnan(left) | np.isnan(right))
    left = left[valid]
    right = right[valid]
    concordant = 0
    discordant = 0
    for i in range(len(left)):
        for j in range(i + 1, len(left)):
            a = np.sign(left[i] - left[j])
            b = np.sign(right[i] - right[j])
            if a == 0 or b == 0:
                continue
            if a == b:
                concordant += 1
            else:
                discordant += 1
    total = concordant + discordant
    return float((concordant - discordant) / total) if total else float("nan")


def load_inputs() -> dict[str, pd.DataFrame]:
    """Load result workbooks used by the judge-driven revision."""
    inputs = {
        "features_norm": pd.read_excel(TABLES_DIR / "appendix1_features_normalized.xlsx"),
        "weights": pd.read_excel(TABLES_DIR / "appendix1_weights_ahp_entropy.xlsx", sheet_name="combined_weights"),
        "entropy_imputation": pd.read_excel(
            TABLES_DIR / "appendix1_weights_ahp_entropy.xlsx", sheet_name="entropy_imputation"
        ),
        "final_ranking": pd.read_excel(TABLES_DIR / "final_problem1_ranking.xlsx", sheet_name="final_ranking"),
        "grade_summary": pd.read_excel(TABLES_DIR / "kmeans_grade_details.xlsx", sheet_name="kmeans_grade_summary"),
        "pairwise": pd.read_excel(TABLES_DIR / "pairwise_comparison_filled.xlsx", sheet_name="pairwise_template"),
        "bt_rank_details": pd.read_excel(TABLES_DIR / "bt_lambda_sensitivity.xlsx", sheet_name="rank_details"),
        "bt_recommendation": pd.read_excel(TABLES_DIR / "bt_lambda_sensitivity.xlsx", sheet_name="recommendation"),
        "bt_sensitivity": pd.read_excel(TABLES_DIR / "bt_lambda_sensitivity.xlsx", sheet_name="summary"),
        "robust_leave_one": pd.read_excel(TABLES_DIR / "problem1_robustness_audit.xlsx", sheet_name="leave_one_indicator"),
    }
    for key in ["features_norm", "final_ranking", "bt_rank_details"]:
        inputs[key] = inputs[key].copy()
        inputs[key]["paper_id"] = inputs[key]["paper_id"].apply(normalize_id)
    inputs["ocr_summary"] = pd.read_excel(TABLES_DIR / "ocr_parse_report.xlsx", sheet_name="summary")
    inputs["ocr_pages"] = pd.read_excel(TABLES_DIR / "ocr_parse_report.xlsx", sheet_name="page_detail")
    return inputs


def feature_columns(features: pd.DataFrame) -> list[str]:
    """Return sorted indicator columns."""
    return sorted([c for c in features.columns if str(c).startswith("I")], key=indicator_number)


def prepare_feature_matrix(features: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Return normalized feature matrix with median imputation for audit recalculation."""
    columns = feature_columns(features)
    matrix = features.sort_values("paper_id")[columns].apply(pd.to_numeric, errors="coerce")
    for column in columns:
        fill_value = matrix[column].median()
        if pd.isna(fill_value):
            fill_value = 0.0
        matrix[column] = matrix[column].fillna(fill_value).clip(0, 1)
    return matrix.reset_index(drop=True), columns


def calculate_topsis_scores(matrix: pd.DataFrame, weights: pd.Series) -> pd.Series:
    """Calculate TOPSIS scores from a normalized feature matrix and weights."""
    weights = weights.reindex(matrix.columns).astype(float)
    weights = weights / weights.sum()
    weighted = matrix.mul(weights, axis=1)
    positive = weighted.max(axis=0)
    negative = weighted.min(axis=0)
    d_plus = np.sqrt(((weighted - positive) ** 2).sum(axis=1))
    d_minus = np.sqrt(((weighted - negative) ** 2).sum(axis=1))
    denom = d_plus + d_minus
    closeness = np.where(denom > 0, d_minus / denom, 0.0)
    return pd.Series(100 * closeness, index=matrix.index)


def combined_weights_for_alpha(weights: pd.DataFrame, columns: list[str], alpha: float) -> pd.Series:
    """Create combined weights under a chosen AHP coefficient alpha."""
    table = weights.set_index("indicator").reindex(columns)
    values = alpha * table["ahp_weight"].astype(float) + (1 - alpha) * table["entropy_weight"].astype(float)
    return values / values.sum()


def final_combined_weights(weights: pd.DataFrame, columns: list[str]) -> pd.Series:
    """Return the final combined weights."""
    table = weights.set_index("indicator").reindex(columns)
    values = table["combined_weight"].astype(float)
    return values / values.sum()


def run_alpha_sensitivity(inputs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Run alpha sensitivity for AHP-entropy weighting."""
    matrix, columns = prepare_feature_matrix(inputs["features_norm"])
    final = inputs["final_ranking"].sort_values("paper_id").reset_index(drop=True)
    baseline_rank = final["rank_final"].astype(int)
    s_bt_scaled = final["S_BT_scaled"].astype(float).reset_index(drop=True)
    rows: list[dict[str, Any]] = []
    for alpha in ALPHA_GRID:
        weights = combined_weights_for_alpha(inputs["weights"], columns, alpha)
        s_base = calculate_topsis_scores(matrix, weights)
        s_rank = LAMBDA_FINAL * s_base + (1 - LAMBDA_FINAL) * s_bt_scaled
        rank = rank_desc(s_rank)
        rank_change = baseline_rank - rank
        rows.append(
            {
                "alpha": alpha,
                "spearman_vs_final_rank": spearman_from_ranks(baseline_rank, rank),
                "max_abs_rank_change": int(rank_change.abs().max()),
                "mean_abs_rank_change": float(rank_change.abs().mean()),
                "top5_same_count": int(
                    len(set(final.loc[baseline_rank <= 5, "paper_id"]) & set(final.loc[rank <= 5, "paper_id"]))
                ),
                "bottom5_same_count": int(
                    len(set(final.loc[baseline_rank >= 26, "paper_id"]) & set(final.loc[rank >= 26, "paper_id"]))
                ),
                "S_rank_min": float(s_rank.min()),
                "S_rank_max": float(s_rank.max()),
            }
        )
    return pd.DataFrame(rows)


def run_removal_scenario(
    matrix: pd.DataFrame,
    weights: pd.Series,
    final: pd.DataFrame,
    removed: list[str],
    scenario: str,
) -> dict[str, Any]:
    """Run one indicator deletion scenario."""
    kept = [column for column in matrix.columns if column not in removed]
    scenario_weights = weights.reindex(kept)
    scenario_weights = scenario_weights / scenario_weights.sum()
    s_base = calculate_topsis_scores(matrix[kept], scenario_weights)
    s_bt_scaled = final["S_BT_scaled"].astype(float).reset_index(drop=True)
    s_rank = LAMBDA_FINAL * s_base + (1 - LAMBDA_FINAL) * s_bt_scaled
    rank = rank_desc(s_rank)
    baseline_rank = final["rank_final"].astype(int)
    rank_change = baseline_rank - rank
    return {
        "scenario": scenario,
        "removed_indicators": ", ".join(removed),
        "removed_count": len(removed),
        "spearman_vs_final_rank": spearman_from_ranks(baseline_rank, rank),
        "max_abs_rank_change": int(rank_change.abs().max()),
        "mean_abs_rank_change": float(rank_change.abs().mean()),
        "top5_same_count": int(
            len(set(final.loc[baseline_rank <= 5, "paper_id"]) & set(final.loc[rank <= 5, "paper_id"]))
        ),
        "bottom5_same_count": int(
            len(set(final.loc[baseline_rank >= 26, "paper_id"]) & set(final.loc[rank >= 26, "paper_id"]))
        ),
    }


def run_indicator_ablation(inputs: dict[str, pd.DataFrame]) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Audit sparse binary and soft indicator sensitivity."""
    matrix, columns = prepare_feature_matrix(inputs["features_norm"])
    final = inputs["final_ranking"].sort_values("paper_id").reset_index(drop=True)
    weights = final_combined_weights(inputs["weights"], columns)
    sparse_binary = []
    binary_like = []
    for column in columns:
        values = set(matrix[column].dropna().round(6).unique())
        zero_count = int((matrix[column] == 0).sum())
        if values <= {0, 1}:
            binary_like.append(column)
            if zero_count >= 10:
                sparse_binary.append(column)
        elif values <= {0, 0.5, 1} and zero_count >= 7:
            sparse_binary.append(column)

    scenarios = [
        ("remove_I16_sensitivity", [c for c in columns if c.startswith("I16_")]),
        ("remove_I04_code_appendix", [c for c in columns if c.startswith("I04_")]),
        (
            "remove_top2_sparse_binary_I16_I04",
            [c for c in columns if c.startswith("I16_") or c.startswith("I04_")],
        ),
        ("remove_all_0_1_binary", binary_like),
        ("remove_sparse_binary_like", sparse_binary),
    ]
    binary_rows = [
        run_removal_scenario(matrix, weights, final, removed, name)
        for name, removed in scenarios
        if removed
    ]

    soft_scenarios = [
        ("remove_I13_method_semantic", [c for c in columns if c.startswith("I13_")]),
        ("remove_I20_innovation", [c for c in columns if c.startswith("I20_")]),
        ("remove_I21_application_value", [c for c in columns if c.startswith("I21_")]),
        (
            "remove_soft_I13_I20_I21",
            [c for c in columns if c.startswith("I13_") or c.startswith("I20_") or c.startswith("I21_")],
        ),
    ]
    soft_rows = [
        run_removal_scenario(matrix, weights, final, removed, name)
        for name, removed in soft_scenarios
        if removed
    ]
    return pd.DataFrame(binary_rows), pd.DataFrame(soft_rows)


def build_bt_independence(inputs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Measure how independent surrogate pairwise decisions are from TOPSIS."""
    pairwise = inputs["pairwise"].copy()
    final_lookup = inputs["final_ranking"].copy()
    final_lookup["paper_id"] = final_lookup["paper_id"].apply(normalize_id)
    sbase = final_lookup.set_index("paper_id")["S_base"].astype(float).to_dict()
    pairwise["paper_i_norm"] = pairwise["paper_i"].apply(normalize_id)
    pairwise["paper_j_norm"] = pairwise["paper_j"].apply(normalize_id)

    directional_rows = []
    signed_products = []
    non_tie_match = []
    for _, row in pairwise.iterrows():
        pid_i = row["paper_i_norm"]
        pid_j = row["paper_j_norm"]
        winner = str(row.get("winner", "")).strip().lower()
        base_diff = sbase[pid_i] - sbase[pid_j]
        base_direction = "i" if base_diff > 0 else "j" if base_diff < 0 else "tie"
        review_sign = 1 if winner == "i" else -1 if winner == "j" else 0
        base_sign = 1 if base_direction == "i" else -1 if base_direction == "j" else 0
        signed_products.append(base_sign * review_sign)
        if winner in {"i", "j"} and base_direction in {"i", "j"}:
            non_tie_match.append(winner == base_direction)
        directional_rows.append(
            {
                "pair_id": row["pair_id"],
                "paper_i": pid_i,
                "paper_j": pid_j,
                "winner_surrogate": winner,
                "winner_by_S_base": base_direction,
                "S_base_gap_i_minus_j": base_diff,
                "direction_match": winner == base_direction if winner in {"i", "j"} else np.nan,
            }
        )

    bt = inputs["bt_rank_details"].copy()
    rank_base = bt.sort_values("paper_id")["rank_base"].astype(int).reset_index(drop=True)
    rank_bt = bt.sort_values("paper_id")["rank_bt"].astype(int).reset_index(drop=True)
    spearman = spearman_from_ranks(rank_base, rank_bt)
    kendall = kendall_tau_from_ranks(rank_base, rank_bt)
    signed_index = float(np.mean(signed_products)) if signed_products else float("nan")
    directional_agreement = float(np.mean(non_tie_match)) if non_tie_match else float("nan")
    final = inputs["final_ranking"]
    rank_change = (final["rank_base"].astype(int) - final["rank_final"].astype(int)).abs()
    rows = [
        {
            "metric": "bt_rank_vs_topsis_spearman",
            "value": spearman,
            "interpretation": "BT潜在排序与TOPSIS基础排序的一致程度；越低说明两路冲突越强。",
        },
        {
            "metric": "bt_rank_vs_topsis_kendall_tau",
            "value": kendall,
            "interpretation": "BT排序与TOPSIS排序的Kendall tau；用于独立性审计。",
        },
        {
            "metric": "pairwise_directional_agreement_excluding_tie",
            "value": directional_agreement,
            "interpretation": "仅比较非tie样本时，surrogate winner与S_base方向一致的比例。",
        },
        {
            "metric": "pairwise_signed_agreement_index",
            "value": signed_index,
            "interpretation": "模板成对方向一致性指数，tie记为0，范围[-1,1]。",
        },
        {
            "metric": "pairwise_independence_index",
            "value": 1 - abs(signed_index),
            "interpretation": "定义为1-|方向一致性指数|；越高说明surrogate与S_base越不重复，但不等价于外部有效性。",
        },
        {
            "metric": "bt_final_max_abs_rank_change",
            "value": float(rank_change.max()),
            "interpretation": "最终采用S_BT_scaled+lambda=0.85后，BT对最终排名的最大影响。",
        },
        {
            "metric": "bt_final_mean_abs_rank_change",
            "value": float(rank_change.mean()),
            "interpretation": "最终采用S_BT_scaled+lambda=0.85后，BT对最终排名的平均影响。",
        },
    ]
    detail = pd.DataFrame(directional_rows)
    summary = pd.DataFrame(rows)
    summary.attrs["detail"] = detail
    return summary


def build_ocr_page_stats(inputs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Summarize OCR page-level character counts."""
    pages = inputs["ocr_pages"]
    rows = []
    for filename, group in pages.groupby(pages.columns[0]):
        chars = pd.to_numeric(group.iloc[:, 2], errors="coerce")
        success = group.iloc[:, 3].astype(bool)
        rows.append(
            {
                "filename": filename,
                "pages": len(group),
                "success_pages": int(success.sum()),
                "failed_pages": int((~success).sum()),
                "total_chars": int(chars.sum()),
                "mean_chars_per_page": float(chars.mean()),
                "median_chars_per_page": float(chars.median()),
                "min_chars_per_page": int(chars.min()),
                "max_chars_per_page": int(chars.max()),
                "min_char_pages": ",".join(map(str, group.loc[chars == chars.min(), group.columns[1]].tolist())),
                "max_char_pages": ",".join(map(str, group.loc[chars == chars.max(), group.columns[1]].tolist())),
            }
        )
    return pd.DataFrame(rows)


def build_missingness_usability(inputs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Explain why partial_missing samples remain usable."""
    features = inputs["features_norm"]
    columns = feature_columns(features)
    missing_count = int(features[columns].isna().sum().sum())
    total_cells = len(features) * len(columns)
    imputed = inputs["entropy_imputation"]["missing_imputed_count"].sum()
    score_confidence = pd.read_excel(TABLES_DIR / "appendix1_topsis_scores.xlsx")
    flag_counts = score_confidence["feature_quality_flag"].value_counts().to_dict()
    confidence_counts = score_confidence["score_confidence"].value_counts().to_dict()
    return pd.DataFrame(
        [
            {"metric": "paper_count", "value": len(features), "note": "附件1样本数。"},
            {"metric": "indicator_count", "value": len(columns), "note": "二级指标数量。"},
            {"metric": "total_feature_cells", "value": total_cells, "note": "论文数×指标数。"},
            {"metric": "missing_feature_cells", "value": missing_count, "note": "标准化特征表中的缺失单元数。"},
            {
                "metric": "missing_feature_rate",
                "value": missing_count / total_cells if total_cells else np.nan,
                "note": "缺失率较低，且仅用于熵权/TOPSIS前的中位数插补。",
            },
            {
                "metric": "entropy_imputed_cells",
                "value": int(imputed),
                "note": "熵权插补报告中的插补单元数。",
            },
            {
                "metric": "feature_quality_flag_counts",
                "value": str(flag_counts),
                "note": "全部为partial_missing时，含义是非致命缺失而非解析失败。",
            },
            {
                "metric": "score_confidence_counts",
                "value": str(confidence_counts),
                "note": "Step 6中无parse_failed样本，均为usable。",
            },
        ]
    )


def build_absolute_guardrail(inputs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Audit simple absolute guardrails without changing final relative grades."""
    features = inputs["features_norm"].copy()
    features["paper_id"] = features["paper_id"].apply(normalize_id)
    final = inputs["final_ranking"][["paper_id", "filename", "rank_final", "grade_final", "S_rank_v2"]].copy()
    merged = final.merge(features, on=["paper_id", "filename"], how="left")

    def col(prefix: str) -> str:
        for column in merged.columns:
            if str(column).startswith(prefix):
                return column
        raise KeyError(prefix)

    c_core = col("I01_")
    c_code = col("I04_")
    c_formula = col("I10_")
    c_variable = col("I11_")
    c_constraint = col("I12_")
    c_result = col("I14_")
    c_sensitivity = col("I16_")
    c_error = col("I17_")

    rows = []
    for _, row in merged.iterrows():
        violations = []
        if float(row[c_core]) < 0.5 and row["grade_final"] in {"优秀", "良好"}:
            violations.append("核心章节完整率<0.5但进入良好及以上")
        if (
            float(row[c_result]) == 0
            and float(row[c_sensitivity]) == 0
            and float(row[c_error]) == 0
            and row["grade_final"] in {"优秀", "良好"}
        ):
            violations.append("结果完整率/灵敏度/误差均为0但进入良好及以上")
        if (
            float(row[c_formula]) < 0.05
            and float(row[c_variable]) <= 0.05
            and float(row[c_constraint]) == 0
            and row["grade_final"] in {"优秀", "良好"}
        ):
            violations.append("公式变量约束显著不足但进入良好及以上")
        rows.append(
            {
                "paper_id": row["paper_id"],
                "filename": row["filename"],
                "rank_final": int(row["rank_final"]),
                "grade_final": row["grade_final"],
                "S_rank_v2": float(row["S_rank_v2"]),
                "core_section_complete": float(row[c_core]),
                "appendix_code": float(row[c_code]),
                "formula_density": float(row[c_formula]),
                "variable_definition": float(row[c_variable]),
                "constraint_completeness": float(row[c_constraint]),
                "result_completeness": float(row[c_result]),
                "sensitivity_exists": float(row[c_sensitivity]),
                "error_analysis_exists": float(row[c_error]),
                "guardrail_violation": bool(violations),
                "violation_reason": "; ".join(violations),
            }
        )
    return pd.DataFrame(rows)


def select_external_anchor_pairs(inputs: dict[str, pd.DataFrame], n_pairs: int = 12) -> pd.DataFrame:
    """Select pairs for real human blind review."""
    pairwise = inputs["pairwise"].copy()
    pairwise["paper_i_norm"] = pairwise["paper_i"].apply(normalize_id)
    pairwise["paper_j_norm"] = pairwise["paper_j"].apply(normalize_id)
    focus = {"02", "07", "25"}
    pairwise["focus_pair"] = pairwise.apply(
        lambda row: int(row["paper_i_norm"] in focus or row["paper_j_norm"] in focus), axis=1
    )
    pairwise["score_gap_abs"] = pd.to_numeric(pairwise["score_gap"], errors="coerce").abs()
    pairwise["rank_gap_abs"] = (
        pd.to_numeric(pairwise["rank_i"], errors="coerce") - pd.to_numeric(pairwise["rank_j"], errors="coerce")
    ).abs()
    pairwise = pairwise.sort_values(["focus_pair", "score_gap_abs", "rank_gap_abs"], ascending=[False, True, True])
    selected = pairwise.head(n_pairs).copy()
    selected["anchor_pair_id"] = [f"A{idx:02d}" for idx in range(1, len(selected) + 1)]
    selected["pair_context"] = selected.apply(
        lambda row: "边界/扫描重点论文对"
        if row["paper_i_norm"] in focus or row["paper_j_norm"] in focus
        else "相邻或接近分数论文对",
        axis=1,
    )
    return selected


def build_external_anchor_template(selected: pd.DataFrame, path: Path) -> Path:
    """Create blank external-anchor templates for real team review."""
    instructions = pd.DataFrame(
        [
            {
                "item": "填写目标",
                "content": "由队员独立盲评8-12对边界论文对，winner只允许i、j、tie；不要参考S_base或surrogate winner。",
            },
            {
                "item": "winner规则",
                "content": "i表示paper_i更好，j表示paper_j更好，tie表示难以区分，空白表示未评。",
            },
            {
                "item": "用途",
                "content": "填完另存为external_anchor_blind_review_filled.xlsx，可用于计算真人盲评与规则化winner一致率。",
            },
            {
                "item": "AHP模板",
                "content": "expert_ahp_primary表用于2-3名队员按1-9标度填写一级指标真实成对偏好，不能由既有权重反推。",
            },
        ]
    )
    blind = selected[
        [
            "anchor_pair_id",
            "pair_context",
            "paper_i_norm",
            "filename_i",
            "paper_j_norm",
            "filename_j",
        ]
    ].rename(columns={"paper_i_norm": "paper_i", "paper_j_norm": "paper_j"})
    blind["reviewer_id"] = ""
    blind["winner"] = ""
    blind["reason"] = ""
    blind["confidence_1_to_5"] = ""

    surrogate_reference = selected[
        ["anchor_pair_id", "pair_id", "paper_i_norm", "paper_j_norm", "winner", "reason", "score_gap"]
    ].rename(
        columns={
            "paper_i_norm": "paper_i",
            "paper_j_norm": "paper_j",
            "winner": "surrogate_winner",
            "reason": "surrogate_reason",
        }
    )
    criteria = ["A1结构规范", "A2逻辑严密", "A3建模质量", "A4结果验证", "A5写作应用"]
    ahp_rows = []
    for reviewer in ["R1", "R2", "R3"]:
        for left, right in combinations(criteria, 2):
            ahp_rows.append(
                {
                    "reviewer_id": reviewer,
                    "criterion_i": left,
                    "criterion_j": right,
                    "winner_criterion": "",
                    "intensity_1_to_9": "",
                    "notes": "",
                }
            )
    ahp = pd.DataFrame(ahp_rows)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        instructions.to_excel(writer, sheet_name="instructions", index=False)
        blind.to_excel(writer, sheet_name="blind_pairwise_review", index=False)
        surrogate_reference.to_excel(writer, sheet_name="surrogate_reference", index=False)
        ahp.to_excel(writer, sheet_name="expert_ahp_primary", index=False)
    format_workbook(path)
    return path


def check_external_anchor_if_filled(path: Path, selected: pd.DataFrame) -> pd.DataFrame:
    """Check real human review file if the team has filled it."""
    filled = TABLES_DIR / "external_anchor_blind_review_filled.xlsx"
    if not filled.exists():
        result = pd.DataFrame(
            [
                {
                    "status": "pending",
                    "filled_file": str(filled),
                    "valid_review_count": 0,
                    "agreement_with_surrogate": np.nan,
                    "matched_count": 0,
                    "mismatched_count": 0,
                    "reviewer_ids": "",
                    "reviewer_id_warning": "",
                    "note": "尚未检测到真人盲评填写文件；不能在论文中声称已完成外部校验。",
                }
            ]
        )
        result.attrs["detail"] = pd.DataFrame()
        return result
    human = pd.read_excel(filled, sheet_name="blind_pairwise_review")
    reference = selected.set_index("anchor_pair_id")["winner"].astype(str).str.lower().to_dict()
    human["winner_norm"] = human["winner"].astype(str).str.strip().str.lower()
    valid = human["winner_norm"].isin(["i", "j", "tie"])
    compared = human.loc[valid].copy()
    compared["surrogate_winner"] = compared["anchor_pair_id"].map(reference)
    compared["is_agree_with_surrogate"] = compared["winner_norm"] == compared["surrogate_winner"]
    if compared.empty:
        agreement = np.nan
        matched_count = 0
    else:
        agreement = float(compared["is_agree_with_surrogate"].mean())
        matched_count = int(compared["is_agree_with_surrogate"].sum())
    reviewer_ids = sorted(
        {
            str(value).strip()
            for value in human.get("reviewer_id", pd.Series(dtype=str)).dropna().tolist()
            if str(value).strip()
        }
    )
    reviewer_id_text = ", ".join(reviewer_ids)
    reviewer_warning = (
        "reviewer_id疑似不是队员编号；若确为队员填写，建议改为R1等编号。"
        if any("ai" in item.lower() for item in reviewer_ids)
        else ""
    )
    status = "completed" if len(compared) >= 8 else "insufficient"
    note = (
        f"已检测到{len(compared)}对有效外部锚复核，一致率{agreement:.2%}。"
        if status == "completed" and not pd.isna(agreement)
        else "有效复核对数不足8对，不建议写入正文。"
    )
    if reviewer_warning:
        note = f"{note}{reviewer_warning}"
    detail_columns = [
        "anchor_pair_id",
        "paper_i",
        "filename_i",
        "paper_j",
        "filename_j",
        "reviewer_id",
        "winner_norm",
        "surrogate_winner",
        "is_agree_with_surrogate",
        "reason",
        "confidence_1_to_5",
    ]
    detail = compared[[column for column in detail_columns if column in compared.columns]].copy()
    result = pd.DataFrame(
        [
            {
                "status": status,
                "filled_file": str(filled),
                "valid_review_count": int(len(compared)),
                "agreement_with_surrogate": agreement,
                "matched_count": matched_count,
                "mismatched_count": int(len(compared) - matched_count),
                "reviewer_ids": reviewer_id_text,
                "reviewer_id_warning": reviewer_warning,
                "note": note,
            }
        ]
    )
    result.attrs["detail"] = detail
    return result


def build_revision_summary(
    missingness: pd.DataFrame,
    ocr_stats: pd.DataFrame,
    bt_independence: pd.DataFrame,
    alpha: pd.DataFrame,
    binary: pd.DataFrame,
    guardrail: pd.DataFrame,
    external_check: pd.DataFrame,
) -> pd.DataFrame:
    """Build a compact revision summary."""
    lookup_missing = dict(zip(missingness["metric"], missingness["value"]))
    lookup_bt = dict(zip(bt_independence["metric"], bt_independence["value"]))
    ocr25 = ocr_stats.loc[ocr_stats["filename"].astype(str).eq("25.pdf")].iloc[0]
    return pd.DataFrame(
        [
            {"revision_item": "OCR页级细节补全", "status": "done", "key_value": f"25.pdf {int(ocr25['success_pages'])}/{int(ocr25['pages'])}页成功，总字数{int(ocr25['total_chars'])}"},
            {"revision_item": "partial_missing可用性说明", "status": "done", "key_value": f"缺失{int(lookup_missing['missing_feature_cells'])}/{int(lookup_missing['total_feature_cells'])}，缺失率{float(lookup_missing['missing_feature_rate']):.4%}"},
            {"revision_item": "BT独立性指数", "status": "done", "key_value": f"independence={float(lookup_bt['pairwise_independence_index']):.6f}"},
            {"revision_item": "AHP alpha敏感性", "status": "done", "key_value": f"min Spearman={alpha['spearman_vs_final_rank'].min():.6f}"},
            {"revision_item": "稀疏二值指标剔除", "status": "done", "key_value": f"top2剔除Spearman={float(binary.loc[binary['scenario']=='remove_top2_sparse_binary_I16_I04','spearman_vs_final_rank'].iloc[0]):.6f}"},
            {"revision_item": "绝对护栏审计", "status": "done", "key_value": f"违反数={int(guardrail['guardrail_violation'].sum())}"},
            {"revision_item": "真人外部锚", "status": str(external_check['status'].iloc[0]), "key_value": str(external_check['note'].iloc[0])},
        ]
    )


def format_workbook(path: Path) -> None:
    """Apply light formatting to an xlsx workbook."""
    if load_workbook is None:
        return
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions
        if PatternFill is not None:
            fill = PatternFill("solid", fgColor="1F4E78")
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(max(max(len(str(cell.value or "")) for cell in column[:200]) + 2, 10), 70)
            sheet.column_dimensions[letter].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(path)


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> Path:
    """Write the Step 11 audit workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    format_workbook(path)
    return path


def dataframe_to_markdown(frame: pd.DataFrame) -> str:
    """Render a small DataFrame as a GitHub-flavored Markdown table."""
    if frame.empty:
        return "_无记录_"
    text_frame = frame.copy()
    for column in text_frame.columns:
        text_frame[column] = text_frame[column].map(
            lambda value: ""
            if pd.isna(value)
            else f"{value:.6f}"
            if isinstance(value, float)
            else str(value)
        )
    headers = [str(column) for column in text_frame.columns]
    rows = text_frame.values.tolist()
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(["---"] * len(headers)) + " |",
    ]
    for row in rows:
        lines.append("| " + " | ".join(str(item) for item in row) + " |")
    return "\n".join(lines)


def write_revision_markdown(
    path: Path,
    summary: pd.DataFrame,
    bt_independence: pd.DataFrame,
    alpha: pd.DataFrame,
    binary: pd.DataFrame,
    guardrail: pd.DataFrame,
    external_template: Path,
) -> Path:
    """Write a concise judge-revision note for paper writing."""
    path.parent.mkdir(parents=True, exist_ok=True)
    bt_lookup = dict(zip(bt_independence["metric"], bt_independence["value"]))
    lines = [
        "# 问题1 Claude评委意见修订记录",
        "",
        "本文件记录按外部评委意见完成的严格修订。所有新增数值均来自既有结果表或本次新增审计脚本，不改动封版的最终排名、得分和等级。",
        "",
        "## 1. 已完成修订",
        "",
        "| 修订项 | 状态 | 关键结果 |",
        "|---|---|---|",
    ]
    for _, row in summary.iterrows():
        lines.append(f"| {row['revision_item']} | {row['status']} | {row['key_value']} |")

    lines += [
        "",
        "## 2. BT重新定位",
        "",
        "BT不再写成主链路校准或人工评分，而定位为“规则化rubric成对比较的一致性与边界风险证据”。",
        f"BT排序与TOPSIS基础排序Spearman为{float(bt_lookup['bt_rank_vs_topsis_spearman']):.6f}，Kendall tau为{float(bt_lookup['bt_rank_vs_topsis_kendall_tau']):.6f}；",
        f"模板成对方向独立性指数为{float(bt_lookup['pairwise_independence_index']):.6f}。该指标只说明两路信号不完全重复，不代表外部有效性。",
        "",
        "## 3. AHP与权重防守",
        "",
        "AHP判断矩阵仍应诚实表述为由预设重要性向量构造的一致偏好矩阵。新增alpha敏感性用于说明组合权重系数变化不会大幅推翻最终排序。",
        "",
        dataframe_to_markdown(alpha),
        "",
        "## 4. 稀疏二值指标防守",
        "",
        "评委指出I16灵敏度分析存在性、I04附录代码存在性权重较高且为稀疏二值指标。本次新增剔除检验，专门检查这些指标是否单独支配最终排序。",
        "",
        dataframe_to_markdown(binary),
        "",
        "## 5. 绝对+相对双轨说明",
        "",
        f"当前最终等级仍是批内相对等级；绝对护栏审计发现违反护栏样本数为{int(guardrail['guardrail_violation'].sum())}。",
        "因此论文中可表述为：本题给出的五级等级是附件1内部相对质量等级，并通过核心章节、结果验证和建模表达护栏检查避免明显不合格样本进入高等级。",
        "",
        "## 6. 外部锚状态",
        "",
    ]
    anchor_status = str(summary.loc[summary["revision_item"] == "真人外部锚", "status"].iloc[0])
    anchor_value = str(summary.loc[summary["revision_item"] == "真人外部锚", "key_value"].iloc[0])
    if anchor_status == "completed":
        lines += [
            f"已检测到外部锚复核填写文件，对应模板为：`{external_template.as_posix()}`。",
            anchor_value,
            "正文可表述为：队员对12对边界论文进行了外部锚复核，并与规则化winner进行一致性比较。",
        ]
    else:
        lines += [
            f"已生成真人盲评与专家AHP模板：`{external_template.as_posix()}`。",
            "在队员未真实填写前，论文不得声称已完成人工外部校验；填写后可运行本脚本自动计算一致率，再补入正文。",
        ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def main() -> None:
    """Run Step 11 judge-driven revision audit."""
    logger = setup_logger(LOGS_DIR / "problem1_judge_revision.log")
    logger.info("Starting Step 11 judge-driven revision audit. Sealed results will not be modified.")
    inputs = load_inputs()

    alpha = run_alpha_sensitivity(inputs)
    binary, soft = run_indicator_ablation(inputs)
    bt_independence = build_bt_independence(inputs)
    bt_pair_detail = bt_independence.attrs["detail"]
    ocr_stats = build_ocr_page_stats(inputs)
    missingness = build_missingness_usability(inputs)
    guardrail = build_absolute_guardrail(inputs)
    selected_pairs = select_external_anchor_pairs(inputs)
    template_path = build_external_anchor_template(
        selected_pairs,
        TABLES_DIR / "external_anchor_blind_review_template.xlsx",
    )
    external_check = check_external_anchor_if_filled(template_path, selected_pairs)
    external_detail = external_check.attrs.get("detail", pd.DataFrame())
    summary = build_revision_summary(missingness, ocr_stats, bt_independence, alpha, binary, guardrail, external_check)

    audit_path = write_workbook(
        TABLES_DIR / "problem1_judge_revision_audit.xlsx",
        {
            "summary": summary,
            "ocr_page_stats": ocr_stats,
            "missingness_usability": missingness,
            "bt_independence": bt_independence,
            "bt_pair_direction_detail": bt_pair_detail,
            "alpha_sensitivity": alpha,
            "binary_indicator_ablation": binary,
            "soft_indicator_ablation": soft,
            "absolute_guardrail_audit": guardrail,
            "external_anchor_pairs": selected_pairs,
            "external_anchor_check": external_check,
            "external_anchor_review_detail": external_detail,
        },
    )
    revision_md = write_revision_markdown(
        PAPER_SECTIONS_DIR / "problem1_judge_revision.md",
        summary,
        bt_independence,
        alpha,
        binary,
        guardrail,
        template_path,
    )

    logger.info("Saved judge revision audit: %s", audit_path)
    logger.info("Saved external anchor template: %s", template_path)
    logger.info("Saved judge revision markdown: %s", revision_md)
    logger.info("Finished Step 11.")

    print("Step 11 judge-driven revision audit finished.")
    print(f"Judge revision audit: {audit_path}")
    print(f"External anchor template: {template_path}")
    print(f"Judge revision markdown: {revision_md}")
    print(f"BT independence index: {float(bt_independence.loc[bt_independence['metric']=='pairwise_independence_index','value'].iloc[0]):.6f}")
    print(f"Alpha sensitivity min Spearman: {float(alpha['spearman_vs_final_rank'].min()):.6f}")
    print(f"Absolute guardrail violations: {int(guardrail['guardrail_violation'].sum())}")
    print(f"External anchor status: {external_check['status'].iloc[0]}")


if __name__ == "__main__":
    main()
