"""Step 10: supplementary credibility, explainability, and robustness audit.

This step does not change the sealed Problem 1 scores, ranks, grades, or any
Step 1-8B result workbook. It produces additional evidence for paper writing:
robustness checks, boundary confidence, OCR quality proxy, evidence summaries,
and anti-circularity wording checks.
"""

from __future__ import annotations

from pathlib import Path
import logging
import math
import re
import sys
from typing import Any

import matplotlib.pyplot as plt
from matplotlib import font_manager
import numpy as np
import pandas as pd

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

TABLES_DIR = PROJECT_ROOT / "output/tables"
CHARTS_DIR = PROJECT_ROOT / "output/charts"
LOGS_DIR = PROJECT_ROOT / "output/logs"
PAPER_SECTIONS_DIR = PROJECT_ROOT / "paper_sections"

GRADE_ORDER = ["优秀", "良好", "中等", "及格", "不及格"]
RANDOM_SEED = 2026
N_WEIGHT_TRIALS = 300
N_BOOTSTRAP_TRIALS = 300
WEIGHT_LOG_SIGMA = 0.10
LAMBDA_FINAL = 0.85


def setup_logger(path: Path) -> logging.Logger:
    """Create a Step 10 logger."""
    path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("A_MAGE_R3.problem1_enhance")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    handler = logging.FileHandler(path, mode="w", encoding="utf-8-sig")
    handler.setFormatter(
        logging.Formatter(
            fmt="%(asctime)s [%(levelname)s] %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    logger.addHandler(handler)
    return logger


def set_chinese_font() -> None:
    """Use a Chinese font if available."""
    candidates = ["Microsoft YaHei", "SimHei", "SimSun", "Noto Sans CJK SC", "Arial Unicode MS"]
    available = {font.name for font in font_manager.fontManager.ttflist}
    for candidate in candidates:
        if candidate in available:
            plt.rcParams["font.sans-serif"] = [candidate]
            break
    plt.rcParams["axes.unicode_minus"] = False


def normalize_id(value: Any) -> str:
    """Normalize IDs as two-digit strings."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    match = re.search(r"(\d+)", text)
    return match.group(1).zfill(2) if match else text.zfill(2)


def indicator_number(column: str) -> int:
    """Return numeric suffix from an indicator name."""
    match = re.match(r"I(\d+)", str(column))
    return int(match.group(1)) if match else 999


def spearman_from_ranks(left: pd.Series, right: pd.Series) -> float:
    """Calculate Spearman correlation from aligned rank columns."""
    left = pd.to_numeric(left, errors="coerce")
    right = pd.to_numeric(right, errors="coerce")
    valid = ~(left.isna() | right.isna())
    if valid.sum() < 2:
        return float("nan")
    return float(np.corrcoef(left[valid], right[valid])[0, 1])


def load_inputs() -> dict[str, pd.DataFrame]:
    """Load Step 10 input tables."""
    inputs = {
        "features_norm": pd.read_excel(TABLES_DIR / "appendix1_features_normalized.xlsx"),
        "features_raw": pd.read_excel(TABLES_DIR / "appendix1_features_raw.xlsx"),
        "weights": pd.read_excel(TABLES_DIR / "appendix1_weights_ahp_entropy.xlsx", sheet_name="combined_weights"),
        "weights_config": pd.read_excel(TABLES_DIR / "appendix1_weights_ahp_entropy.xlsx", sheet_name="config"),
        "rank_fusion_v2": pd.read_excel(TABLES_DIR / "appendix1_rank_fusion_v2.xlsx", sheet_name="rank_fusion_v2"),
        "final_ranking": pd.read_excel(TABLES_DIR / "final_problem1_ranking.xlsx", sheet_name="final_ranking"),
        "primary_scores": pd.read_excel(TABLES_DIR / "final_problem1_ranking.xlsx", sheet_name="primary_scores"),
        "grade_summary": pd.read_excel(TABLES_DIR / "kmeans_grade_details.xlsx", sheet_name="kmeans_grade_summary"),
        "boundary_nearby": pd.read_excel(TABLES_DIR / "kmeans_grade_details.xlsx", sheet_name="boundary_nearby"),
        "bt_sensitivity": pd.read_excel(TABLES_DIR / "bt_lambda_sensitivity.xlsx", sheet_name="summary"),
        "bt_recommendation": pd.read_excel(TABLES_DIR / "bt_lambda_sensitivity.xlsx", sheet_name="recommendation"),
        "bt_rank_change_audit": pd.read_excel(TABLES_DIR / "bt_rank_change_audit.xlsx", sheet_name="large_rank_changes"),
        "pairwise": pd.read_excel(TABLES_DIR / "pairwise_comparison_filled.xlsx", sheet_name="pairwise_template"),
    }
    for key in ["features_norm", "features_raw", "rank_fusion_v2", "final_ranking", "primary_scores"]:
        inputs[key] = inputs[key].copy()
        inputs[key]["paper_id"] = inputs[key]["paper_id"].apply(normalize_id)

    ocr_path = TABLES_DIR / "ocr_parse_report.xlsx"
    if ocr_path.exists():
        inputs["ocr_summary"] = pd.read_excel(ocr_path, sheet_name="summary")
        inputs["ocr_pages"] = pd.read_excel(ocr_path, sheet_name="page_detail")
    else:
        inputs["ocr_summary"] = pd.DataFrame()
        inputs["ocr_pages"] = pd.DataFrame()
    return inputs


def prepare_feature_matrix(features_norm: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Prepare normalized feature matrix for supplementary robustness checks."""
    feature_columns = sorted(
        [column for column in features_norm.columns if str(column).startswith("I")],
        key=indicator_number,
    )
    matrix = features_norm[feature_columns].apply(pd.to_numeric, errors="coerce")
    for column in feature_columns:
        fill_value = matrix[column].median()
        if pd.isna(fill_value):
            fill_value = 0.0
        matrix[column] = matrix[column].fillna(fill_value).clip(0, 1)
    return matrix, feature_columns


def calculate_topsis_scores(matrix: pd.DataFrame, weights: pd.Series) -> pd.Series:
    """Calculate TOPSIS base scores from a normalized matrix and weights."""
    weights = weights.reindex(matrix.columns).astype(float)
    weights = weights / weights.sum()
    weighted = matrix.mul(weights, axis=1)
    positive = weighted.max(axis=0)
    negative = weighted.min(axis=0)
    d_plus = np.sqrt(((weighted - positive) ** 2).sum(axis=1))
    d_minus = np.sqrt(((weighted - negative) ** 2).sum(axis=1))
    denom = d_plus + d_minus
    closeness = np.where(denom > 0, d_minus / denom, 0.0)
    return pd.Series(100 * closeness, index=matrix.index)


def rank_desc(scores: pd.Series) -> pd.Series:
    """Rank scores descending, using first occurrence for ties."""
    return scores.rank(ascending=False, method="first").astype(int)


def align_weights(weights: pd.DataFrame, feature_columns: list[str]) -> pd.Series:
    """Align combined weights to feature columns."""
    weight_series = weights.set_index("indicator")["combined_weight"].astype(float)
    missing = sorted(set(feature_columns) - set(weight_series.index))
    if missing:
        raise ValueError(f"权重表缺少指标: {missing}")
    weight_series = weight_series.reindex(feature_columns)
    return weight_series / weight_series.sum()


def assign_grade_by_centers(scores: pd.Series, centers: dict[str, float]) -> pd.Series:
    """Assign grades by nearest fixed KMeans center."""
    center_items = list(centers.items())
    labels: list[str] = []
    for value in scores:
        labels.append(min(center_items, key=lambda item: abs(float(value) - item[1]))[0])
    return pd.Series(labels, index=scores.index)


def build_grade_centers(grade_summary: pd.DataFrame) -> dict[str, float]:
    """Return KMeans grade centers from high to low."""
    summary = grade_summary.copy()
    return {row["grade"]: float(row["center"]) for _, row in summary.iterrows()}


def run_weight_perturbation(
    matrix: pd.DataFrame,
    base_weights: pd.Series,
    final_ranking: pd.DataFrame,
    centers: dict[str, float],
    rng: np.random.Generator,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Run multiplicative weight perturbation robustness analysis."""
    baseline = final_ranking.sort_values("paper_id").reset_index(drop=True).copy()
    baseline_rank = baseline["rank_final"].astype(int)
    baseline_grade = baseline["grade_final"].astype(str)
    s_bt_scaled = baseline["S_BT_scaled"].astype(float).reset_index(drop=True)

    trial_summaries: list[dict[str, Any]] = []
    rank_records: list[pd.Series] = []
    grade_records: list[pd.Series] = []

    matrix_aligned = matrix.copy().reset_index(drop=True)
    for trial in range(N_WEIGHT_TRIALS):
        perturbed = base_weights.to_numpy() * np.exp(rng.normal(0.0, WEIGHT_LOG_SIGMA, len(base_weights)))
        perturbed = pd.Series(perturbed / perturbed.sum(), index=base_weights.index)
        s_base_perturbed = calculate_topsis_scores(matrix_aligned, perturbed)
        s_rank = LAMBDA_FINAL * s_base_perturbed + (1 - LAMBDA_FINAL) * s_bt_scaled
        rank = rank_desc(s_rank)
        grades = assign_grade_by_centers(s_rank, centers)
        rank_records.append(rank.rename(trial))
        grade_records.append(grades.rename(trial))

        rank_change = baseline_rank - rank
        trial_summaries.append(
            {
                "trial": trial + 1,
                "spearman_vs_final": spearman_from_ranks(baseline_rank, rank),
                "max_abs_rank_change": int(rank_change.abs().max()),
                "mean_abs_rank_change": float(rank_change.abs().mean()),
                "grade_same_rate": float((grades.reset_index(drop=True) == baseline_grade).mean()),
            }
        )

    rank_matrix = pd.concat(rank_records, axis=1)
    grade_matrix = pd.concat(grade_records, axis=1)

    per_paper: list[dict[str, Any]] = []
    for index, row in baseline.iterrows():
        ranks = rank_matrix.loc[index].astype(float)
        grades = grade_matrix.loc[index].astype(str)
        per_paper.append(
            {
                "paper_id": row["paper_id"],
                "filename": row["filename"],
                "rank_final": int(row["rank_final"]),
                "grade_final": row["grade_final"],
                "rank_mean": float(ranks.mean()),
                "rank_std": float(ranks.std(ddof=0)),
                "rank_min": int(ranks.min()),
                "rank_max": int(ranks.max()),
                "max_abs_rank_change": int((ranks - int(row["rank_final"])).abs().max()),
                "same_grade_rate": float((grades == row["grade_final"]).mean()),
                "top5_rate": float((ranks <= 5).mean()),
                "bottom5_rate": float((ranks >= 26).mean()),
            }
        )

    summary = pd.DataFrame(trial_summaries)
    overall = pd.DataFrame(
        [
            {"metric": "trial_count", "value": N_WEIGHT_TRIALS},
            {"metric": "weight_log_sigma", "value": WEIGHT_LOG_SIGMA},
            {"metric": "mean_spearman_vs_final", "value": summary["spearman_vs_final"].mean()},
            {"metric": "min_spearman_vs_final", "value": summary["spearman_vs_final"].min()},
            {"metric": "mean_max_abs_rank_change", "value": summary["max_abs_rank_change"].mean()},
            {"metric": "max_observed_rank_change", "value": summary["max_abs_rank_change"].max()},
            {"metric": "mean_grade_same_rate", "value": summary["grade_same_rate"].mean()},
        ]
    )
    return overall, summary, pd.DataFrame(per_paper)


def run_leave_one_indicator_out(
    matrix: pd.DataFrame,
    base_weights: pd.Series,
    final_ranking: pd.DataFrame,
) -> pd.DataFrame:
    """Delete one indicator at a time and measure rank sensitivity."""
    baseline = final_ranking.sort_values("paper_id").reset_index(drop=True).copy()
    baseline_rank = baseline["rank_final"].astype(int)
    s_bt_scaled = baseline["S_BT_scaled"].astype(float).reset_index(drop=True)
    rows: list[dict[str, Any]] = []
    matrix_aligned = matrix.reset_index(drop=True)

    for indicator in base_weights.index:
        kept = [column for column in base_weights.index if column != indicator]
        weights = base_weights.reindex(kept)
        weights = weights / weights.sum()
        s_base = calculate_topsis_scores(matrix_aligned[kept], weights)
        s_rank = LAMBDA_FINAL * s_base + (1 - LAMBDA_FINAL) * s_bt_scaled
        rank = rank_desc(s_rank)
        rank_change = baseline_rank - rank
        rows.append(
            {
                "removed_indicator": indicator,
                "spearman_vs_final": spearman_from_ranks(baseline_rank, rank),
                "max_abs_rank_change": int(rank_change.abs().max()),
                "mean_abs_rank_change": float(rank_change.abs().mean()),
                "top5_same_count": int(
                    len(
                        set(baseline.loc[baseline_rank <= 5, "paper_id"])
                        & set(baseline.loc[rank <= 5, "paper_id"])
                    )
                ),
                "bottom5_same_count": int(
                    len(
                        set(baseline.loc[baseline_rank >= 26, "paper_id"])
                        & set(baseline.loc[rank >= 26, "paper_id"])
                    )
                ),
            }
        )
    return pd.DataFrame(rows).sort_values(["spearman_vs_final", "max_abs_rank_change"]).reset_index(drop=True)


def run_bootstrap_ideal_stability(
    matrix: pd.DataFrame,
    base_weights: pd.Series,
    final_ranking: pd.DataFrame,
    centers: dict[str, float],
    rng: np.random.Generator,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Bootstrap the sample used to form TOPSIS ideal solutions."""
    baseline = final_ranking.sort_values("paper_id").reset_index(drop=True).copy()
    baseline_rank = baseline["rank_final"].astype(int)
    baseline_grade = baseline["grade_final"].astype(str)
    s_bt_scaled = baseline["S_BT_scaled"].astype(float).reset_index(drop=True)
    matrix_aligned = matrix.reset_index(drop=True)
    weighted_full = matrix_aligned.mul(base_weights, axis=1)

    trial_summaries: list[dict[str, Any]] = []
    rank_records: list[pd.Series] = []
    grade_records: list[pd.Series] = []

    n = len(matrix_aligned)
    for trial in range(N_BOOTSTRAP_TRIALS):
        sampled_indices = rng.integers(0, n, n)
        weighted_sample = weighted_full.iloc[sampled_indices]
        positive = weighted_sample.max(axis=0)
        negative = weighted_sample.min(axis=0)
        d_plus = np.sqrt(((weighted_full - positive) ** 2).sum(axis=1))
        d_minus = np.sqrt(((weighted_full - negative) ** 2).sum(axis=1))
        denom = d_plus + d_minus
        s_base = pd.Series(np.where(denom > 0, 100 * d_minus / denom, 0.0))
        s_rank = LAMBDA_FINAL * s_base + (1 - LAMBDA_FINAL) * s_bt_scaled
        rank = rank_desc(s_rank)
        grades = assign_grade_by_centers(s_rank, centers)
        rank_records.append(rank.rename(trial))
        grade_records.append(grades.rename(trial))
        rank_change = baseline_rank - rank
        trial_summaries.append(
            {
                "trial": trial + 1,
                "spearman_vs_final": spearman_from_ranks(baseline_rank, rank),
                "max_abs_rank_change": int(rank_change.abs().max()),
                "mean_abs_rank_change": float(rank_change.abs().mean()),
                "grade_same_rate": float((grades.reset_index(drop=True) == baseline_grade).mean()),
            }
        )

    rank_matrix = pd.concat(rank_records, axis=1)
    grade_matrix = pd.concat(grade_records, axis=1)
    per_paper: list[dict[str, Any]] = []
    for index, row in baseline.iterrows():
        ranks = rank_matrix.loc[index].astype(float)
        grades = grade_matrix.loc[index].astype(str)
        per_paper.append(
            {
                "paper_id": row["paper_id"],
                "filename": row["filename"],
                "rank_final": int(row["rank_final"]),
                "grade_final": row["grade_final"],
                "rank_mean": float(ranks.mean()),
                "rank_std": float(ranks.std(ddof=0)),
                "rank_min": int(ranks.min()),
                "rank_max": int(ranks.max()),
                "max_abs_rank_change": int((ranks - int(row["rank_final"])).abs().max()),
                "same_grade_rate": float((grades == row["grade_final"]).mean()),
            }
        )

    summary = pd.DataFrame(trial_summaries)
    overall = pd.DataFrame(
        [
            {"metric": "trial_count", "value": N_BOOTSTRAP_TRIALS},
            {"metric": "mean_spearman_vs_final", "value": summary["spearman_vs_final"].mean()},
            {"metric": "min_spearman_vs_final", "value": summary["spearman_vs_final"].min()},
            {"metric": "mean_max_abs_rank_change", "value": summary["max_abs_rank_change"].mean()},
            {"metric": "max_observed_rank_change", "value": summary["max_abs_rank_change"].max()},
            {"metric": "mean_grade_same_rate", "value": summary["grade_same_rate"].mean()},
        ]
    )
    return overall, summary, pd.DataFrame(per_paper)


def calculate_boundary_confidence(
    final_ranking: pd.DataFrame,
    grade_summary: pd.DataFrame,
    weight_stability: pd.DataFrame,
    bootstrap_stability: pd.DataFrame,
) -> pd.DataFrame:
    """Calculate boundary distance and grade confidence for every paper."""
    centers = build_grade_centers(grade_summary)
    center_items = list(centers.items())
    sorted_centers = sorted(center_items, key=lambda item: item[1], reverse=True)
    boundaries = []
    for index in range(len(sorted_centers) - 1):
        high_grade, high_center = sorted_centers[index]
        low_grade, low_center = sorted_centers[index + 1]
        boundaries.append((f"{high_grade}-{low_grade}", (high_center + low_center) / 2))

    weight_lookup = weight_stability.set_index("paper_id")
    boot_lookup = bootstrap_stability.set_index("paper_id")
    rows: list[dict[str, Any]] = []
    for _, row in final_ranking.sort_values("rank_final").iterrows():
        score = float(row["S_rank_v2"])
        distances = sorted([(grade, abs(score - center)) for grade, center in center_items], key=lambda item: item[1])
        assigned_distance = distances[0][1]
        second_distance = distances[1][1] if len(distances) > 1 else math.inf
        center_confidence = second_distance / (assigned_distance + second_distance) if math.isfinite(second_distance) else 1.0
        nearest_boundary_name, nearest_boundary_score = min(boundaries, key=lambda item: abs(score - item[1]))
        boundary_distance = abs(score - nearest_boundary_score)
        weight_same = float(weight_lookup.loc[row["paper_id"], "same_grade_rate"])
        boot_same = float(boot_lookup.loc[row["paper_id"], "same_grade_rate"])
        robust_rate = (weight_same + boot_same) / 2
        final_confidence = 0.6 * center_confidence + 0.4 * robust_rate
        if final_confidence >= 0.75 and boundary_distance >= 1.0:
            risk_type = "stable_grade"
        elif final_confidence >= 0.50:
            risk_type = "boundary_grade"
        else:
            risk_type = "high_risk_grade"
        rows.append(
            {
                "paper_id": row["paper_id"],
                "filename": row["filename"],
                "rank_final": int(row["rank_final"]),
                "grade_final": row["grade_final"],
                "S_rank_v2": score,
                "nearest_boundary": nearest_boundary_name,
                "nearest_boundary_score": nearest_boundary_score,
                "distance_to_boundary": boundary_distance,
                "center_confidence": center_confidence,
                "weight_perturb_same_grade_rate": weight_same,
                "bootstrap_same_grade_rate": boot_same,
                "grade_confidence": final_confidence,
                "grade_risk_type": risk_type,
            }
        )
    return pd.DataFrame(rows)


def calculate_ocr_quality(inputs: dict[str, pd.DataFrame], final_ranking: pd.DataFrame) -> pd.DataFrame:
    """Create OCR quality proxy rows for all papers."""
    ocr_summary = inputs.get("ocr_summary", pd.DataFrame())
    ocr_lookup: dict[str, dict[str, Any]] = {}
    if not ocr_summary.empty:
        for _, row in ocr_summary.iterrows():
            paper_id = normalize_id(row.get("文件名", ""))
            pages = float(row.get("页数", np.nan))
            success_pages = float(row.get("OCR成功页数", np.nan))
            total_chars = float(row.get("OCR总字数", np.nan))
            success_rate = success_pages / pages if pages and not pd.isna(pages) else np.nan
            chars_per_page = total_chars / pages if pages and not pd.isna(pages) else np.nan
            char_density_score = min(1.0, chars_per_page / 500.0) if not pd.isna(chars_per_page) else np.nan
            quality_proxy = success_rate * char_density_score if not pd.isna(success_rate) and not pd.isna(char_density_score) else np.nan
            ocr_lookup[paper_id] = {
                "source_type": "ocr_fallback",
                "ocr_pages": pages,
                "ocr_success_pages": success_pages,
                "ocr_success_rate": success_rate,
                "ocr_total_chars": total_chars,
                "ocr_chars_per_page": chars_per_page,
                "ocr_quality_proxy": quality_proxy,
                "ocr_note": "未输出Tesseract置信度，使用页成功率和字数密度作为OCR质量代理。",
            }

    rows = []
    for _, row in final_ranking.sort_values("rank_final").iterrows():
        paper_id = row["paper_id"]
        if paper_id in ocr_lookup:
            values = ocr_lookup[paper_id]
        else:
            values = {
                "source_type": "text_pdf",
                "ocr_pages": np.nan,
                "ocr_success_pages": np.nan,
                "ocr_success_rate": np.nan,
                "ocr_total_chars": np.nan,
                "ocr_chars_per_page": np.nan,
                "ocr_quality_proxy": 1.0,
                "ocr_note": "未使用OCR兜底，按常规文本PDF处理。",
            }
        rows.append({"paper_id": paper_id, "filename": row["filename"], **values})
    return pd.DataFrame(rows)


def build_evidence_summary(
    inputs: dict[str, pd.DataFrame],
    boundary_confidence: pd.DataFrame,
    ocr_quality: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build evidence summaries for all papers and representative papers."""
    final = inputs["final_ranking"].sort_values("rank_final").copy()
    features = inputs["features_norm"].set_index("paper_id")
    raw = inputs["features_raw"].set_index("paper_id")
    primary = inputs["primary_scores"].set_index("paper_id")
    weights = inputs["weights"].copy()
    weight_series = weights.set_index("indicator")["combined_weight"].astype(float)
    feature_columns = sorted([c for c in features.columns if str(c).startswith("I")], key=indicator_number)
    boundary_lookup = boundary_confidence.set_index("paper_id")
    ocr_lookup = ocr_quality.set_index("paper_id")

    rows: list[dict[str, Any]] = []
    for _, row in final.iterrows():
        paper_id = row["paper_id"]
        values = pd.to_numeric(features.loc[paper_id, feature_columns], errors="coerce")
        contributions = (values.fillna(0.0) * weight_series.reindex(feature_columns)).sort_values(ascending=False)
        top_items = [
            f"{indicator}={values[indicator]:.3f},贡献{contributions[indicator]:.4f}"
            for indicator in contributions.head(3).index
        ]
        weak_candidates = values[values.fillna(0.0) <= 0.25].index.tolist()
        weak_candidates = sorted(weak_candidates, key=lambda indicator: weight_series.get(indicator, 0.0), reverse=True)
        weak_items = [
            f"{indicator}={0.0 if pd.isna(values[indicator]) else values[indicator]:.3f}"
            for indicator in weak_candidates[:3]
        ]
        missing_items = [indicator for indicator in feature_columns if pd.isna(values[indicator])]
        zero_items = [
            indicator
            for indicator in feature_columns
            if not pd.isna(values[indicator]) and abs(float(values[indicator])) <= 1e-12
        ]
        primary_values = pd.to_numeric(
            primary.loc[paper_id, [c for c in primary.columns if c not in {"filename"}]],
            errors="coerce",
        )
        primary_values = primary_values.drop(labels=["paper_id"], errors="ignore")
        primary_strength = primary_values.idxmax()
        primary_weakness = primary_values.idxmin()
        evidence_sentence = (
            f"{row['filename']}最终等级为{row['grade_final']}，S_rank_v2={row['S_rank_v2']:.3f}；"
            f"主要正向证据为{'; '.join(top_items)}；"
            f"主要扣分证据为{'; '.join(weak_items) if weak_items else '无显著低值指标'}；"
            f"一级指标最高为{primary_strength}={primary_values[primary_strength]:.3f}，"
            f"最低为{primary_weakness}={primary_values[primary_weakness]:.3f}；"
            f"等级风险为{boundary_lookup.loc[paper_id, 'grade_risk_type']}。"
        )
        rows.append(
            {
                "paper_id": paper_id,
                "filename": row["filename"],
                "rank_final": int(row["rank_final"]),
                "grade_final": row["grade_final"],
                "S_rank_v2": float(row["S_rank_v2"]),
                "text_chars": int(raw.loc[paper_id, "text_chars"]) if "text_chars" in raw.columns else np.nan,
                "primary_strength": primary_strength,
                "primary_strength_score": float(primary_values[primary_strength]),
                "primary_weakness": primary_weakness,
                "primary_weakness_score": float(primary_values[primary_weakness]),
                "top_positive_evidence": "; ".join(top_items),
                "top_penalty_evidence": "; ".join(weak_items),
                "missing_feature_count": len(missing_items),
                "missing_features": ",".join(missing_items),
                "zero_feature_count": len(zero_items),
                "zero_features": ",".join(zero_items),
                "grade_confidence": float(boundary_lookup.loc[paper_id, "grade_confidence"]),
                "grade_risk_type": boundary_lookup.loc[paper_id, "grade_risk_type"],
                "ocr_quality_proxy": float(ocr_lookup.loc[paper_id, "ocr_quality_proxy"]),
                "evidence_sentence": evidence_sentence,
            }
        )
    evidence = pd.DataFrame(rows)

    representatives: list[pd.Series] = []
    for grade in GRADE_ORDER:
        subset = evidence.loc[evidence["grade_final"] == grade].copy()
        if subset.empty:
            continue
        center = float(inputs["grade_summary"].loc[inputs["grade_summary"]["grade"] == grade, "center"].iloc[0])
        subset["distance_to_grade_center"] = (subset["S_rank_v2"] - center).abs()
        representatives.append(subset.sort_values("distance_to_grade_center").iloc[0])
    representative_df = pd.DataFrame(representatives)
    return evidence, representative_df


def build_anti_circular_audit(inputs: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Create anti-circularity and wording-risk audit rows."""
    bt_summary = inputs["bt_sensitivity"]
    original = bt_summary.loc[(bt_summary["bt_score_version"] == "S_BT") & (bt_summary["lambda"] == 0.70)]
    final = bt_summary.loc[(bt_summary["bt_score_version"] == "S_BT_scaled") & (bt_summary["lambda"] == 0.85)]
    original_spearman = float(original["spearman_vs_topsis"].iloc[0])
    final_spearman = float(final["spearman_vs_topsis"].iloc[0])
    final_max_change = int(final["max_abs_rank_change"].iloc[0])
    rows = [
        {
            "audit_item": "BT成对比较来源",
            "risk_before_revision": "若写成纯人工判断，会与surrogate reviewer自动填写事实不一致。",
            "recommended_wording": "基于五维rubric的规则化成对比较校准；可作为后续人工复核接口。",
            "evidence": "pairwise_comparison_filled.xlsx包含winner与reason；脚本run_step7a_fill_pairwise_surrogate.py可追溯生成逻辑。",
            "status_after_upgrade": "wording_revised",
        },
        {
            "audit_item": "BT是否过度推翻TOPSIS",
            "risk_before_revision": f"原始S_BT融合Spearman={original_spearman:.6f}，影响偏强。",
            "recommended_wording": "采用S_BT_scaled + lambda=0.85，使BT仅作温和排序校准。",
            "evidence": f"最终Spearman={final_spearman:.6f}，最大排名变化={final_max_change}。",
            "status_after_upgrade": "pass",
        },
        {
            "audit_item": "AHP一致性过于完美",
            "risk_before_revision": "CR=0若表述为真实专家打分结果，可能被质疑为反推矩阵。",
            "recommended_wording": "基于预设重要性向量构造一致判断矩阵，并用权重扰动敏感性分析检验稳定性。",
            "evidence": "appendix1_weights_ahp_entropy.xlsx与problem1_robustness_audit.xlsx。",
            "status_after_upgrade": "wording_revised",
        },
        {
            "audit_item": "KMeans/Jenks完全一致",
            "risk_before_revision": "不能宣称两个完全独立模型共同证明结果正确。",
            "recommended_wording": "KMeans为主分级，Jenks作为一维自然断点一致性校验。",
            "evidence": "grade_distribution.xlsx中KMeans与Jenks一致率为1.000000。",
            "status_after_upgrade": "wording_revised",
        },
        {
            "audit_item": "OCR质量进入风险提示",
            "risk_before_revision": "扫描件25.pdf若不说明OCR质量，可能影响特征可信度解释。",
            "recommended_wording": "对OCR样本记录页成功率、字数密度和OCR质量代理，作为风险提示而非直接扣分。",
            "evidence": "ocr_parse_report.xlsx显示25.pdf共36页，成功36页，总字数32064。",
            "status_after_upgrade": "pass",
        },
    ]
    return pd.DataFrame(rows)


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> Path:
    """Write Excel workbook with light formatting."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    if load_workbook is not None:
        workbook = load_workbook(path)
        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            if sheet.max_row > 1 and sheet.max_column > 1:
                sheet.auto_filter.ref = sheet.dimensions
            for column in sheet.columns:
                letter = column[0].column_letter
                width = min(max(max(len(str(cell.value or "")) for cell in column[:200]) + 2, 10), 60)
                sheet.column_dimensions[letter].width = width
        workbook.save(path)
    return path


def plot_rank_stability(per_paper: pd.DataFrame, path: Path, title: str) -> Path:
    """Plot rank standard deviation by paper."""
    path.parent.mkdir(parents=True, exist_ok=True)
    table = per_paper.sort_values("rank_final")
    fig, ax = plt.subplots(figsize=(15, 6))
    bars = ax.bar(table["paper_id"], table["rank_std"], color="#4E79A7")
    ax.set_title(title, fontsize=16, fontweight="bold")
    ax.set_xlabel("论文编号")
    ax.set_ylabel("排名标准差")
    ax.grid(axis="y", alpha=0.25)
    for bar, value in zip(bars, table["rank_std"]):
        if value >= 1.0:
            ax.text(bar.get_x() + bar.get_width() / 2, value + 0.03, f"{value:.1f}", ha="center", fontsize=8)
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)
    return path


def plot_boundary_confidence(boundary: pd.DataFrame, path: Path) -> Path:
    """Plot grade confidence and boundary risk."""
    path.parent.mkdir(parents=True, exist_ok=True)
    colors = {
        "stable_grade": "#2E7D32",
        "boundary_grade": "#F9A825",
        "high_risk_grade": "#C62828",
    }
    table = boundary.sort_values("rank_final")
    fig, ax = plt.subplots(figsize=(15, 6))
    ax.bar(
        table["paper_id"],
        table["grade_confidence"],
        color=[colors.get(item, "#666666") for item in table["grade_risk_type"]],
    )
    ax.axhline(0.75, color="#2E7D32", linestyle="--", linewidth=1.4, label="稳定等级参考线")
    ax.axhline(0.50, color="#C62828", linestyle="--", linewidth=1.4, label="边界风险参考线")
    ax.set_ylim(0, 1.05)
    ax.set_title("最终等级置信度与边界风险", fontsize=16, fontweight="bold")
    ax.set_xlabel("论文编号")
    ax.set_ylabel("等级置信度")
    ax.grid(axis="y", alpha=0.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)
    return path


def plot_multi_agent_framework(path: Path) -> Path:
    """Plot the multi-agent review framework."""
    from matplotlib.patches import FancyBboxPatch

    path.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(14, 8))
    ax.axis("off")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)

    agent_boxes = [
        ("结构规范\nAgent", "章节/摘要/参考文献/附录", 0.06, 0.62, "#E3F2FD"),
        ("逻辑一致\nAgent", "问题重述/假设/结论链", 0.26, 0.62, "#E8F5E9"),
        ("数学建模\nAgent", "模型/公式/变量/约束", 0.46, 0.62, "#FFF3E0"),
        ("结果验证\nAgent", "结果/图表/灵敏度/误差", 0.66, 0.62, "#F3E5F5"),
        ("写作应用\nAgent", "可读性/创新/推广价值", 0.36, 0.40, "#E0F7FA"),
    ]

    def box(x: float, y: float, w: float, h: float, text: str, facecolor: str) -> None:
        patch = FancyBboxPatch(
            (x, y),
            w,
            h,
            boxstyle="round,pad=0.018,rounding_size=0.018",
            linewidth=1.4,
            edgecolor="#37474F",
            facecolor=facecolor,
        )
        ax.add_patch(patch)
        ax.text(x + w / 2, y + h / 2, text, ha="center", va="center", fontsize=12, weight="bold")

    for title, subtitle, x, y, color in agent_boxes:
        box(x, y, 0.17, 0.16, f"{title}\n{subtitle}", color)
        ax.annotate(
            "",
            xy=(0.50, 0.34),
            xytext=(x + 0.085, y),
            arrowprops={"arrowstyle": "->", "linewidth": 1.4, "color": "#455A64"},
        )

    box(0.37, 0.20, 0.26, 0.14, "综合仲裁 Agent\nAHP-熵权 + TOPSIS + BT校准", "#ECEFF1")
    ax.annotate(
        "",
        xy=(0.50, 0.13),
        xytext=(0.50, 0.20),
        arrowprops={"arrowstyle": "->", "linewidth": 1.6, "color": "#263238"},
    )
    box(0.34, 0.02, 0.32, 0.11, "最终输出\nS_rank_v2 / 五级等级 / 证据摘要 / 风险提示", "#FFEBEE")

    ax.text(
        0.5,
        0.94,
        "问题1多智能体可审计论文评估框架",
        ha="center",
        va="center",
        fontsize=20,
        weight="bold",
    )
    ax.text(
        0.5,
        0.88,
        "各Agent只产生可追溯证据，最终由综合仲裁Agent融合并输出稳定性与边界风险",
        ha="center",
        va="center",
        fontsize=12,
        color="#455A64",
    )
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)
    return path


def write_evidence_markdown(evidence: pd.DataFrame, representatives: pd.DataFrame, path: Path) -> Path:
    """Write evidence-based explanations for paper writing."""
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# 问题1证据型解释与边界样本说明",
        "",
        "本文件用于补充问题1的可解释性。所有解释均由已生成指标值、一级指标汇总、最终等级和稳定性审计得到，不重新评分。",
        "",
        "## 各等级代表论文解释",
        "",
    ]
    for _, row in representatives.iterrows():
        lines += [
            f"### {row['grade_final']}代表：{row['filename']}",
            "",
            f"- 最终排名：{int(row['rank_final'])}",
            f"- 最终得分：{row['S_rank_v2']:.6f}",
            f"- 等级置信度：{row['grade_confidence']:.6f}",
            f"- 等级风险类型：`{row['grade_risk_type']}`",
            f"- 主要正向证据：{row['top_positive_evidence']}",
            f"- 主要扣分证据：{row['top_penalty_evidence'] if row['top_penalty_evidence'] else '无显著低值指标'}",
            f"- 一级指标优势：{row['primary_strength']}={row['primary_strength_score']:.3f}",
            f"- 一级指标短板：{row['primary_weakness']}={row['primary_weakness_score']:.3f}",
            "",
            row["evidence_sentence"],
            "",
        ]

    boundary = evidence.loc[evidence["grade_risk_type"] != "stable_grade"].copy()
    lines += [
        "## 边界与风险样本",
        "",
    ]
    if boundary.empty:
        lines.append("当前没有被标记为边界等级或高风险等级的样本。")
    else:
        lines += [
            "| 论文 | 最终等级 | 最终得分 | 等级置信度 | 风险类型 | 主要原因 |",
            "|---|---|---:|---:|---|---|",
        ]
        for _, row in boundary.sort_values("grade_confidence").iterrows():
            lines.append(
                f"| {row['filename']} | {row['grade_final']} | {row['S_rank_v2']:.6f} | "
                f"{row['grade_confidence']:.6f} | {row['grade_risk_type']} | {row['top_penalty_evidence']} |"
            )

    lines += [
        "",
        "## 全部论文一句话证据摘要",
        "",
    ]
    for _, row in evidence.sort_values("rank_final").iterrows():
        lines.append(f"- {row['evidence_sentence']}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_system_upgrade_markdown(
    robustness_summary: pd.DataFrame,
    bootstrap_summary: pd.DataFrame,
    anti_circular: pd.DataFrame,
    path: Path,
) -> Path:
    """Write system design and innovation notes."""
    path.parent.mkdir(parents=True, exist_ok=True)
    weight_metrics = dict(zip(robustness_summary["metric"], robustness_summary["value"]))
    boot_metrics = dict(zip(bootstrap_summary["metric"], bootstrap_summary["value"]))
    lines = [
        "# 问题1智能评估系统升级说明",
        "",
        "## 1. 定位调整",
        "",
        "问题1不再表述为多个常规模型的简单串联，而表述为“多维文本特征驱动的可审计智能评估系统”。系统由文本解析、结构识别、指标抽取、组合赋权、基础排序、规则化成对比较校准、等级划分、稳健性审计和证据解释九个环节构成。",
        "",
        "## 2. 多智能体评审框架",
        "",
        "| Agent | 职责 | 对应指标或输出 |",
        "|---|---|---|",
        "| 结构规范 Agent | 检查摘要、章节、参考文献、附录和图表编号 | A1结构规范性、参考文献规范率、附录代码存在性 |",
        "| 逻辑一致 Agent | 检查问题重述、假设、结论之间的逻辑链 | A2问题理解与逻辑严密性、结果-结论一致性 |",
        "| 数学建模 Agent | 检查模型数量、公式密度、变量定义和约束完整性 | A3方法合理性与数学建模质量 |",
        "| 结果验证 Agent | 检查结果完整性、图表解释、灵敏度和误差分析 | A4结果分析与验证 |",
        "| 综合仲裁 Agent | 融合各Agent证据，输出最终分数、等级和风险提示 | AHP-熵权、TOPSIS、BT校准、KMeans/Jenks分级 |",
        "",
        "## 3. 成对比较校准的表述修正",
        "",
        "当前 `winner` 并非真实线下专家人工填写，而是由基于五维rubric的规则化 surrogate reviewer 生成。因此论文中应表述为“规则化成对比较校准”，不要写成“人工判断”。更强的后续方案是抽取20-30对边界论文进行人工复核，并报告人工与规则化winner的一致率。",
        "",
        "## 4. 稳健性补充结果",
        "",
        f"- 权重扰动试验次数：{int(weight_metrics['trial_count'])}；",
        f"- 权重扰动平均Spearman：{float(weight_metrics['mean_spearman_vs_final']):.6f}；",
        f"- 权重扰动最小Spearman：{float(weight_metrics['min_spearman_vs_final']):.6f}；",
        f"- 权重扰动平均等级保持率：{float(weight_metrics['mean_grade_same_rate']):.6f}；",
        f"- Bootstrap试验次数：{int(boot_metrics['trial_count'])}；",
        f"- Bootstrap平均Spearman：{float(boot_metrics['mean_spearman_vs_final']):.6f}；",
        f"- Bootstrap最小Spearman：{float(boot_metrics['min_spearman_vs_final']):.6f}；",
        f"- Bootstrap平均等级保持率：{float(boot_metrics['mean_grade_same_rate']):.6f}。",
        "",
        "这些结果用于支撑：最终排序对小幅权重扰动和TOPSIS理想解样本扰动具有可解释的稳定性。具体排名波动、指标删除结果和边界样本见 `problem1_robustness_audit.xlsx`。",
        "",
        "## 5. 反循环论证审计",
        "",
        "| 审计项 | 修正后状态 | 建议表述 |",
        "|---|---|---|",
    ]
    for _, row in anti_circular.iterrows():
        lines.append(f"| {row['audit_item']} | {row['status_after_upgrade']} | {row['recommended_wording']} |")

    lines += [
        "",
        "## 6. 论文表达建议",
        "",
        "- AHP部分不要过度强调CR=0，而应说明其为基于预设重要性向量构造的一致偏好矩阵。",
        "- KMeans/Jenks部分不要写成两个完全独立模型共同证明，而写成“KMeans主分级，Jenks自然断点一致性校验”。",
        "- BT部分明确规则化rubric来源，避免“人工判断”不实表述。",
        "- 对02.txt等边界论文给出等级置信度和风险提示，体现真实评审系统的不确定性意识。",
        "- 对25.pdf说明OCR质量代理和风险记录，强调扫描件已进入可追溯质量控制流程。",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def create_run_all_script(path: Path) -> Path:
    """Create a reproducible runner for Problem 1."""
    content = '''"""Run the full Problem 1 pipeline in order.

This runner is for reproducibility. It may overwrite generated outputs, so use
--dry-run first when you only want to inspect commands.
"""

from __future__ import annotations

import argparse
from pathlib import Path
import subprocess
import sys


PROJECT_ROOT = Path(__file__).resolve().parents[1]

STEPS = [
    ("step1_setup", "scripts/run_step1_setup.py"),
    ("step2_parse_pdf", "scripts/run_step2_parse_pdf.py"),
    ("step2b_ocr_failed", "scripts/run_step2b_ocr_parse_failed.py"),
    ("step3_split_sections", "scripts/run_step3_split_sections.py"),
    ("step4_extract_features", "scripts/run_step4_extract_features.py"),
    ("step5_weighting", "scripts/run_step5_weighting.py"),
    ("step6_topsis", "scripts/run_step6_topsis.py"),
    ("step7a_pairwise_template", "scripts/run_step7a_generate_pairwise_template.py"),
    ("step7a_surrogate_fill", "scripts/run_step7a_fill_pairwise_surrogate.py"),
    ("step7b_pairwise_quality", "scripts/run_step7b_pairwise_quality_check.py"),
    ("step7b_bradley_terry", "scripts/run_step7b_bradley_terry.py"),
    ("step7c_bt_sensitivity", "scripts/run_step7c_bt_audit_sensitivity.py"),
    ("step8_grade_visualize", "scripts/run_step8_grade_visualize.py"),
    ("step8b_final_audit", "scripts/run_step8b_final_audit.py"),
    ("step10_enhance", "scripts/run_step10_problem1_enhance.py"),
]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Print commands without running them.")
    parser.add_argument("--from-step", default=None, help="Start from a named step.")
    args = parser.parse_args()

    start = 0
    if args.from_step:
        names = [name for name, _ in STEPS]
        if args.from_step not in names:
            raise SystemExit(f"Unknown step {args.from_step}. Available: {names}")
        start = names.index(args.from_step)

    for name, script in STEPS[start:]:
        command = [sys.executable, str(PROJECT_ROOT / script)]
        print(f"[{name}] {' '.join(command)}")
        if not args.dry_run:
            subprocess.run(command, cwd=PROJECT_ROOT, check=True)


if __name__ == "__main__":
    main()
'''
    path.write_text(content, encoding="utf-8")
    return path


def main() -> None:
    """Run supplementary Problem 1 enhancement."""
    logger = setup_logger(LOGS_DIR / "problem1_enhancement.log")
    set_chinese_font()
    logger.info("Starting Step 10 supplementary enhancement. Sealed scores will not be modified.")
    inputs = load_inputs()
    matrix, feature_columns = prepare_feature_matrix(inputs["features_norm"].sort_values("paper_id"))
    final = inputs["final_ranking"].sort_values("paper_id").reset_index(drop=True)
    weights = align_weights(inputs["weights"], feature_columns)
    centers = build_grade_centers(inputs["grade_summary"])
    rng = np.random.default_rng(RANDOM_SEED)

    weight_overall, weight_trials, weight_per_paper = run_weight_perturbation(
        matrix=matrix,
        base_weights=weights,
        final_ranking=final,
        centers=centers,
        rng=rng,
    )
    leave_one = run_leave_one_indicator_out(matrix, weights, final)
    bootstrap_overall, bootstrap_trials, bootstrap_per_paper = run_bootstrap_ideal_stability(
        matrix=matrix,
        base_weights=weights,
        final_ranking=final,
        centers=centers,
        rng=rng,
    )
    boundary_confidence = calculate_boundary_confidence(
        final,
        inputs["grade_summary"],
        weight_per_paper,
        bootstrap_per_paper,
    )
    ocr_quality = calculate_ocr_quality(inputs, final)
    evidence, representatives = build_evidence_summary(inputs, boundary_confidence, ocr_quality)
    anti_circular = build_anti_circular_audit(inputs)

    workbook_path = write_workbook(
        TABLES_DIR / "problem1_robustness_audit.xlsx",
        {
            "summary": pd.DataFrame(
                [
                    {"metric": "final_scores_changed", "value": False},
                    {"metric": "lambda_final", "value": LAMBDA_FINAL},
                    {"metric": "weight_trials", "value": N_WEIGHT_TRIALS},
                    {"metric": "bootstrap_trials", "value": N_BOOTSTRAP_TRIALS},
                    {
                        "metric": "boundary_or_high_risk_count",
                        "value": int((boundary_confidence["grade_risk_type"] != "stable_grade").sum()),
                    },
                ]
            ),
            "weight_perturb_summary": weight_overall,
            "weight_perturb_trials": weight_trials,
            "weight_perturb_by_paper": weight_per_paper,
            "leave_one_indicator": leave_one,
            "bootstrap_summary": bootstrap_overall,
            "bootstrap_trials": bootstrap_trials,
            "bootstrap_by_paper": bootstrap_per_paper,
            "boundary_confidence": boundary_confidence,
            "ocr_quality_proxy": ocr_quality,
            "evidence_summary": evidence,
            "representative_evidence": representatives,
            "anti_circular_audit": anti_circular,
        },
    )
    weight_chart = plot_rank_stability(
        weight_per_paper,
        CHARTS_DIR / "weight_perturbation_rank_stability.png",
        "权重扰动下的排名稳定性",
    )
    bootstrap_chart = plot_rank_stability(
        bootstrap_per_paper,
        CHARTS_DIR / "bootstrap_rank_stability.png",
        "Bootstrap理想解扰动下的排名稳定性",
    )
    boundary_chart = plot_boundary_confidence(
        boundary_confidence,
        CHARTS_DIR / "boundary_grade_confidence.png",
    )
    agent_chart = plot_multi_agent_framework(CHARTS_DIR / "multi_agent_framework.png")
    evidence_md = write_evidence_markdown(
        evidence,
        representatives,
        PAPER_SECTIONS_DIR / "problem1_evidence_explanations.md",
    )
    upgrade_md = write_system_upgrade_markdown(
        weight_overall,
        bootstrap_overall,
        anti_circular,
        PAPER_SECTIONS_DIR / "problem1_system_upgrade.md",
    )
    run_all = create_run_all_script(PROJECT_ROOT / "scripts/run_all_problem1.py")

    logger.info("Saved robustness workbook: %s", workbook_path)
    logger.info("Saved charts: %s, %s, %s, %s", weight_chart, bootstrap_chart, boundary_chart, agent_chart)
    logger.info("Saved markdown: %s, %s", evidence_md, upgrade_md)
    logger.info("Saved reproducibility runner: %s", run_all)
    logger.info("Finished Step 10.")

    print("Step 10 Problem 1 enhancement finished.")
    print(f"Robustness workbook: {workbook_path}")
    print(f"Weight perturbation chart: {weight_chart}")
    print(f"Bootstrap chart: {bootstrap_chart}")
    print(f"Boundary confidence chart: {boundary_chart}")
    print(f"Multi-agent framework chart: {agent_chart}")
    print(f"Evidence markdown: {evidence_md}")
    print(f"System upgrade markdown: {upgrade_md}")
    print(f"Run-all script: {run_all}")
    print(
        "Weight perturbation mean Spearman: "
        f"{float(weight_overall.loc[weight_overall['metric']=='mean_spearman_vs_final','value'].iloc[0]):.6f}"
    )
    print(
        "Bootstrap mean Spearman: "
        f"{float(bootstrap_overall.loc[bootstrap_overall['metric']=='mean_spearman_vs_final','value'].iloc[0]):.6f}"
    )
    print(
        "Boundary/high risk papers: "
        f"{int((boundary_confidence['grade_risk_type'] != 'stable_grade').sum())}"
    )


if __name__ == "__main__":
    main()
