"""Pairwise-comparison utilities for Bradley-Terry preparation.

Step 7A only generates a human annotation template. It does not estimate
Bradley-Terry parameters, alter TOPSIS scores, or classify papers into grades.
"""

from __future__ import annotations

from collections import Counter
from pathlib import Path
import logging
import re
from typing import Any

import numpy as np
import pandas as pd


LOGGER_NAME = "A_MAGE_R3.pairwise_template"
QUALITY_LOGGER_NAME = "A_MAGE_R3.pairwise_quality_check"
BT_LOGGER_NAME = "A_MAGE_R3.bradley_terry"

TEMPLATE_COLUMNS = [
    "pair_id",
    "paper_i",
    "filename_i",
    "score_i",
    "rank_i",
    "paper_j",
    "filename_j",
    "score_j",
    "rank_j",
    "score_gap",
    "winner",
    "reason",
]
ALLOWED_WINNERS = {"i", "j", "tie", ""}


def setup_pairwise_template_logger(log_path: Path) -> logging.Logger:
    """Configure the Step 7A template logger."""
    log_path = Path(log_path)
    log_path.parent.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger(LOGGER_NAME)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger


def _get_logger() -> logging.Logger:
    """Return the Step 7A logger."""
    return logging.getLogger(LOGGER_NAME)


def setup_pairwise_quality_logger(log_path: Path) -> logging.Logger:
    """Configure the pairwise input quality-check logger."""
    log_path = Path(log_path)
    log_path.parent.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger(QUALITY_LOGGER_NAME)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger


def setup_bradley_terry_logger(log_path: Path) -> logging.Logger:
    """Configure the Bradley-Terry fitting logger."""
    log_path = Path(log_path)
    log_path.parent.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger(BT_LOGGER_NAME)
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setLevel(logging.INFO)
    formatter = logging.Formatter(
        fmt="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger


def _normalize_id(value: Any) -> str:
    """Normalize paper IDs to two-digit strings."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    match = re.search(r"(\d+)", text)
    return match.group(1).zfill(2) if match else text


def load_topsis_scores(score_path: Path) -> pd.DataFrame:
    """Load and validate TOPSIS base scores."""
    score_path = Path(score_path)
    try:
        scores = pd.read_excel(score_path, sheet_name="topsis_scores")
    except ValueError:
        scores = pd.read_excel(score_path)

    required = {"paper_id", "filename", "S_base", "rank_base"}
    missing = required - set(scores.columns)
    if missing:
        raise ValueError(f"TOPSIS 评分表缺少必要列: {sorted(missing)}")

    scores = scores.copy()
    scores["paper_id"] = scores["paper_id"].apply(_normalize_id)
    scores["S_base"] = pd.to_numeric(scores["S_base"], errors="coerce")
    scores["rank_base"] = pd.to_numeric(scores["rank_base"], errors="coerce")
    if scores["S_base"].isna().any() or scores["rank_base"].isna().any():
        raise ValueError("TOPSIS 评分表中 S_base 或 rank_base 存在非数值/缺失。")

    return scores.sort_values(["rank_base", "paper_id"]).reset_index(drop=True)


def select_pairwise_comparisons(
    scores: pd.DataFrame,
    close_gap_threshold: float = 3.0,
    boundary_scores: list[float] | tuple[float, ...] = (60.0, 50.0, 40.0),
    min_appearances_per_paper: int = 2,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Select paper pairs for human comparison without assigning winners."""
    scores = scores.sort_values(["rank_base", "paper_id"]).reset_index(drop=True)
    pair_sources: dict[tuple[str, str], set[str]] = {}
    pair_indices: dict[tuple[str, str], tuple[int, int]] = {}

    def add_pair(index_a: int, index_b: int, source: str) -> None:
        if index_a == index_b:
            return
        rank_a = float(scores.loc[index_a, "rank_base"])
        rank_b = float(scores.loc[index_b, "rank_base"])
        high_index, low_index = (index_a, index_b) if rank_a < rank_b else (index_b, index_a)
        paper_a = str(scores.loc[high_index, "paper_id"])
        paper_b = str(scores.loc[low_index, "paper_id"])
        key = (paper_a, paper_b)
        pair_sources.setdefault(key, set()).add(source)
        pair_indices[key] = (high_index, low_index)

    n = len(scores)
    if n < 2:
        raise ValueError("至少需要 2 篇论文才能生成成对比较模板。")

    for index in range(n - 1):
        add_pair(index, index + 1, "adjacent_rank")

    for index_i in range(n):
        score_i = float(scores.loc[index_i, "S_base"])
        for index_j in range(index_i + 1, n):
            score_j = float(scores.loc[index_j, "S_base"])
            gap = abs(score_i - score_j)
            if gap <= close_gap_threshold:
                add_pair(index_i, index_j, f"score_gap_le_{close_gap_threshold:g}")
            elif scores.loc[index_i, "rank_base"] < scores.loc[index_j, "rank_base"] and score_i >= score_j:
                # Scores are sorted descending, so further rows only get farther.
                break

    for boundary in boundary_scores:
        candidates: list[tuple[float, int, int]] = []
        for index in range(n - 1):
            score_i = float(scores.loc[index, "S_base"])
            score_j = float(scores.loc[index + 1, "S_base"])
            if score_i >= boundary >= score_j:
                midpoint_distance = abs(((score_i + score_j) / 2) - boundary)
                gap = abs(score_i - score_j)
                candidates.append((midpoint_distance + 0.001 * gap, index, index + 1))
        if candidates:
            _, index_i, index_j = sorted(candidates, key=lambda item: item[0])[0]
            add_pair(index_i, index_j, f"boundary_near_{boundary:g}")

    def appearance_counts() -> Counter[str]:
        counts: Counter[str] = Counter()
        for paper_i, paper_j in pair_sources:
            counts[paper_i] += 1
            counts[paper_j] += 1
        for paper_id in scores["paper_id"]:
            counts.setdefault(str(paper_id), 0)
        return counts

    progress = True
    while progress:
        progress = False
        counts = appearance_counts()
        undercovered = [paper_id for paper_id, count in counts.items() if count < min_appearances_per_paper]
        if not undercovered:
            break

        rank_lookup = {str(row.paper_id): int(pos) for pos, row in enumerate(scores.itertuples(index=False))}
        for paper_id in sorted(undercovered, key=lambda item: rank_lookup[item]):
            index = rank_lookup[paper_id]
            candidates = []
            for other_index in range(n):
                if other_index == index:
                    continue
                rank_distance = abs(other_index - index)
                score_gap = abs(float(scores.loc[index, "S_base"]) - float(scores.loc[other_index, "S_base"]))
                candidates.append((rank_distance, score_gap, other_index))
            for _, _, other_index in sorted(candidates):
                before = len(pair_sources)
                add_pair(index, other_index, "coverage_min_appearances")
                if len(pair_sources) > before:
                    progress = True
                    break

    rows: list[dict[str, Any]] = []
    selection_rows: list[dict[str, Any]] = []
    sorted_pairs = sorted(
        pair_sources,
        key=lambda key: (
            min(scores.loc[pair_indices[key][0], "rank_base"], scores.loc[pair_indices[key][1], "rank_base"]),
            abs(float(scores.loc[pair_indices[key][0], "S_base"]) - float(scores.loc[pair_indices[key][1], "S_base"])),
            key,
        ),
    )
    for pair_number, key in enumerate(sorted_pairs, start=1):
        index_i, index_j = pair_indices[key]
        row_i = scores.loc[index_i]
        row_j = scores.loc[index_j]
        score_i = float(row_i["S_base"])
        score_j = float(row_j["S_base"])
        pair_id = f"P{pair_number:03d}"
        rows.append(
            {
                "pair_id": pair_id,
                "paper_i": str(row_i["paper_id"]),
                "filename_i": row_i["filename"],
                "score_i": score_i,
                "rank_i": int(row_i["rank_base"]),
                "paper_j": str(row_j["paper_id"]),
                "filename_j": row_j["filename"],
                "score_j": score_j,
                "rank_j": int(row_j["rank_base"]),
                "score_gap": abs(score_i - score_j),
                "winner": "",
                "reason": "",
            }
        )
        selection_rows.append(
            {
                "pair_id": pair_id,
                "selection_reasons": ",".join(sorted(pair_sources[key])),
            }
        )

    template = pd.DataFrame(rows, columns=TEMPLATE_COLUMNS)
    counts = appearance_counts()
    coverage = pd.DataFrame(
        [
            {
                "paper_id": paper_id,
                "filename": scores.loc[scores["paper_id"].eq(paper_id), "filename"].iloc[0],
                "rank_base": int(scores.loc[scores["paper_id"].eq(paper_id), "rank_base"].iloc[0]),
                "S_base": float(scores.loc[scores["paper_id"].eq(paper_id), "S_base"].iloc[0]),
                "appearance_count": int(counts[paper_id]),
                "is_under_minimum": int(counts[paper_id]) < min_appearances_per_paper,
            }
            for paper_id in scores["paper_id"]
        ]
    ).sort_values(["appearance_count", "rank_base"]).reset_index(drop=True)
    selection_notes = pd.DataFrame(selection_rows)
    return template, coverage, selection_notes


def _add_excel_validation(output_path: Path) -> None:
    """Add a winner drop-down list and light formatting to the workbook."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception:
        return

    workbook = load_workbook(output_path)
    sheet = workbook["pairwise_template"]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    winner_col = TEMPLATE_COLUMNS.index("winner") + 1
    winner_letter = sheet.cell(row=1, column=winner_col).column_letter
    validation = DataValidation(type="list", formula1='"i,j,tie"', allow_blank=True)
    validation.error = "winner 只能填写 i、j、tie，或留空。"
    validation.prompt = "填 i 表示 paper_i 更好；填 j 表示 paper_j 更好；填 tie 表示难以区分。"
    sheet.add_data_validation(validation)
    validation.add(f"{winner_letter}2:{winner_letter}{sheet.max_row}")

    widths = {
        "A": 10,
        "B": 10,
        "C": 14,
        "D": 12,
        "E": 9,
        "F": 10,
        "G": 14,
        "H": 12,
        "I": 9,
        "J": 12,
        "K": 12,
        "L": 36,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for worksheet in workbook.worksheets:
        if worksheet.title != "pairwise_template":
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(output_path)


def generate_pairwise_template(
    score_path: Path,
    output_path: Path,
    log_path: Path | None = None,
    close_gap_threshold: float = 3.0,
    boundary_scores: list[float] | tuple[float, ...] = (60.0, 50.0, 40.0),
    min_appearances_per_paper: int = 2,
) -> dict[str, Any]:
    """Generate a blank human pairwise-comparison template."""
    logger = setup_pairwise_template_logger(log_path) if log_path is not None else _get_logger()
    score_path = Path(score_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Starting Step 7A pairwise template generation: score_path=%s", score_path)
    scores = load_topsis_scores(score_path)
    template, coverage, selection_notes = select_pairwise_comparisons(
        scores=scores,
        close_gap_threshold=close_gap_threshold,
        boundary_scores=boundary_scores,
        min_appearances_per_paper=min_appearances_per_paper,
    )

    rules = pd.DataFrame(
        [
            {"key": "input_score_file", "value": str(score_path)},
            {"key": "selection_rule_1", "value": "all adjacent rank pairs"},
            {"key": "selection_rule_2", "value": f"all pairs with S_base gap <= {close_gap_threshold:g}"},
            {"key": "selection_rule_3", "value": f"closest pairs around score boundaries {list(boundary_scores)}"},
            {"key": "selection_rule_4", "value": f"coverage fill until each paper appears at least {min_appearances_per_paper} times where possible"},
            {"key": "winner_rule_i", "value": "winner=i means paper_i is judged better by human reviewer"},
            {"key": "winner_rule_j", "value": "winner=j means paper_j is judged better by human reviewer"},
            {"key": "winner_rule_tie", "value": "winner=tie means quality is hard to distinguish"},
            {"key": "blank_rule", "value": "blank winner means not yet compared"},
            {"key": "no_auto_winner", "value": "TOPSIS S_base is only used to choose pairs, never to fill winner"},
        ]
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        template.to_excel(writer, index=False, sheet_name="pairwise_template")
        coverage.to_excel(writer, index=False, sheet_name="coverage_summary")
        selection_notes.to_excel(writer, index=False, sheet_name="selection_notes")
        rules.to_excel(writer, index=False, sheet_name="instructions")

    _add_excel_validation(output_path)

    insufficient = coverage.loc[coverage["is_under_minimum"], "paper_id"].astype(str).tolist()
    logger.info("Pairwise template saved: %s", output_path)
    logger.info("Pairs generated: %s", len(template))
    logger.info("Average appearances per paper: %.4f", float(coverage["appearance_count"].mean()))
    logger.info("Under-covered papers: %s", ",".join(insufficient) if insufficient else "none")
    logger.info("Winner and reason columns are intentionally blank.")
    logger.info("Finished Step 7A pairwise template generation.")
    return {
        "template": template,
        "coverage": coverage,
        "selection_notes": selection_notes,
        "rules": rules,
        "output_path": output_path,
        "insufficient_papers": insufficient,
    }


def _normalize_winner(value: Any) -> str:
    """Normalize a winner cell for validation."""
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def load_pairwise_comparisons(pairwise_path: Path) -> pd.DataFrame:
    """Load the filled pairwise comparison worksheet."""
    pairwise_path = Path(pairwise_path)
    try:
        comparisons = pd.read_excel(pairwise_path, sheet_name="pairwise_template")
    except ValueError:
        comparisons = pd.read_excel(pairwise_path)

    required = set(TEMPLATE_COLUMNS)
    missing = required - set(comparisons.columns)
    if missing:
        raise ValueError(f"成对比较表缺少必要列: {sorted(missing)}")

    comparisons = comparisons.copy()
    comparisons["paper_i"] = comparisons["paper_i"].apply(_normalize_id)
    comparisons["paper_j"] = comparisons["paper_j"].apply(_normalize_id)
    comparisons["winner_normalized"] = comparisons["winner"].apply(_normalize_winner)
    return comparisons


def _paper_universe(comparisons: pd.DataFrame) -> list[str]:
    """Return all paper IDs appearing in the comparison table."""
    papers = sorted(set(comparisons["paper_i"].astype(str)) | set(comparisons["paper_j"].astype(str)))
    return [paper for paper in papers if paper]


def validate_winner_column(comparisons: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Validate winner cells and return annotated comparisons plus invalid rows."""
    checked = comparisons.copy()
    checked["is_blank_winner"] = checked["winner_normalized"].eq("")
    checked["is_legal_winner"] = checked["winner_normalized"].isin(ALLOWED_WINNERS)
    checked["is_valid_comparison"] = checked["winner_normalized"].isin({"i", "j", "tie"})
    invalid = checked.loc[~checked["is_legal_winner"], TEMPLATE_COLUMNS + ["winner_normalized"]].copy()
    return checked, invalid


def calculate_valid_participation(comparisons: pd.DataFrame) -> pd.DataFrame:
    """Count each paper's valid pairwise-comparison appearances."""
    papers = _paper_universe(comparisons)
    valid = comparisons.loc[comparisons["is_valid_comparison"]].copy()
    rows: list[dict[str, Any]] = []
    for paper_id in papers:
        as_i = valid["paper_i"].eq(paper_id)
        as_j = valid["paper_j"].eq(paper_id)
        total = int((as_i | as_j).sum())
        wins = int(((as_i & valid["winner_normalized"].eq("i")) | (as_j & valid["winner_normalized"].eq("j"))).sum())
        losses = int(((as_i & valid["winner_normalized"].eq("j")) | (as_j & valid["winner_normalized"].eq("i"))).sum())
        ties = int(((as_i | as_j) & valid["winner_normalized"].eq("tie")).sum())
        filename = ""
        match_i = comparisons.loc[comparisons["paper_i"].eq(paper_id), "filename_i"]
        match_j = comparisons.loc[comparisons["paper_j"].eq(paper_id), "filename_j"]
        if len(match_i):
            filename = str(match_i.iloc[0])
        elif len(match_j):
            filename = str(match_j.iloc[0])
        rows.append(
            {
                "paper_id": paper_id,
                "filename": filename,
                "valid_comparison_count": total,
                "win_count": wins,
                "loss_count": losses,
                "tie_count": ties,
                "is_under_minimum": total < 2,
            }
        )
    return pd.DataFrame(rows).sort_values(["valid_comparison_count", "paper_id"]).reset_index(drop=True)


def connected_components(comparisons: pd.DataFrame) -> tuple[list[list[str]], pd.DataFrame]:
    """Calculate undirected connected components from valid comparisons."""
    papers = _paper_universe(comparisons)
    adjacency: dict[str, set[str]] = {paper: set() for paper in papers}
    for _, row in comparisons.loc[comparisons["is_valid_comparison"]].iterrows():
        paper_i = str(row["paper_i"])
        paper_j = str(row["paper_j"])
        if not paper_i or not paper_j or paper_i == paper_j:
            continue
        adjacency.setdefault(paper_i, set()).add(paper_j)
        adjacency.setdefault(paper_j, set()).add(paper_i)

    components: list[list[str]] = []
    visited: set[str] = set()
    for paper in papers:
        if paper in visited:
            continue
        stack = [paper]
        current: list[str] = []
        visited.add(paper)
        while stack:
            node = stack.pop()
            current.append(node)
            for neighbor in sorted(adjacency.get(node, set())):
                if neighbor not in visited:
                    visited.add(neighbor)
                    stack.append(neighbor)
        components.append(sorted(current))

    component_rows = []
    for index, component in enumerate(sorted(components, key=lambda item: (-len(item), item)), start=1):
        component_rows.append(
            {
                "component_id": index,
                "component_size": len(component),
                "paper_ids": ",".join(component),
            }
        )
    return components, pd.DataFrame(component_rows)


def suggest_pairs_to_connect_components(components: list[list[str]]) -> pd.DataFrame:
    """Suggest minimal paper pairs if the comparison graph is disconnected."""
    if len(components) <= 1:
        return pd.DataFrame(columns=["suggested_pair_id", "paper_i", "paper_j", "reason"])

    ordered = sorted(components, key=lambda item: (-len(item), item))
    base = ordered[0]
    rows = []
    for index, component in enumerate(ordered[1:], start=1):
        rows.append(
            {
                "suggested_pair_id": f"S{index:03d}",
                "paper_i": base[0],
                "paper_j": component[0],
                "reason": "连接不同连通分量，保证 Bradley-Terry 输入图连通",
            }
        )
    return pd.DataFrame(rows)


def check_pairwise_quality(
    pairwise_path: Path,
    output_path: Path,
    log_path: Path | None = None,
    minimum_valid_comparisons_per_paper: int = 2,
) -> dict[str, Any]:
    """Check pairwise comparison data quality without fitting BT."""
    logger = setup_pairwise_quality_logger(log_path) if log_path is not None else logging.getLogger(QUALITY_LOGGER_NAME)
    pairwise_path = Path(pairwise_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Starting pairwise quality check: %s", pairwise_path)
    comparisons = load_pairwise_comparisons(pairwise_path)
    checked, invalid_winners = validate_winner_column(comparisons)
    participation = calculate_valid_participation(checked)
    components, component_table = connected_components(checked)
    suggestions = suggest_pairs_to_connect_components(components)

    valid_count = int(checked["is_valid_comparison"].sum())
    tie_count = int(checked["winner_normalized"].eq("tie").sum())
    blank_count = int(checked["is_blank_winner"].sum())
    invalid_count = int((~checked["is_legal_winner"]).sum())
    undercovered = participation.loc[
        participation["valid_comparison_count"] < minimum_valid_comparisons_per_paper,
        "paper_id",
    ].astype(str).tolist()
    is_connected = len(components) == 1

    summary_items = [
        ("total_pairs", len(checked)),
        ("valid_comparison_count", valid_count),
        ("tie_count", tie_count),
        ("blank_winner_count", blank_count),
        ("invalid_winner_count", invalid_count),
        ("paper_count", len(participation)),
        ("minimum_valid_comparisons_per_paper", minimum_valid_comparisons_per_paper),
        ("undercovered_paper_count", len(undercovered)),
        ("undercovered_papers", ",".join(undercovered) if undercovered else "none"),
        ("comparison_graph_connected", is_connected),
        ("connected_component_count", len(components)),
        ("can_enter_bt_fitting", invalid_count == 0 and blank_count == 0 and not undercovered and is_connected),
    ]
    summary = pd.DataFrame(
        [
            {
                "item": item,
                "value": "TRUE" if value is True else "FALSE" if value is False else str(value),
            }
            for item, value in summary_items
        ]
    )

    validation_rows = checked[
        [
            "pair_id",
            "paper_i",
            "paper_j",
            "winner",
            "winner_normalized",
            "is_blank_winner",
            "is_legal_winner",
            "is_valid_comparison",
        ]
    ].copy()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="summary")
        validation_rows.to_excel(writer, index=False, sheet_name="winner_validation")
        invalid_winners.to_excel(writer, index=False, sheet_name="invalid_winners")
        participation.to_excel(writer, index=False, sheet_name="paper_participation")
        component_table.to_excel(writer, index=False, sheet_name="graph_components")
        suggestions.to_excel(writer, index=False, sheet_name="suggested_pairs")

    logger.info("Total pairs=%s valid=%s tie=%s blank=%s invalid=%s", len(checked), valid_count, tie_count, blank_count, invalid_count)
    logger.info("Undercovered papers: %s", ",".join(undercovered) if undercovered else "none")
    logger.info("Comparison graph connected=%s components=%s", is_connected, len(components))
    if not is_connected:
        logger.warning("Comparison graph is disconnected; suggested pairs saved to suggested_pairs sheet.")
    logger.info("Pairwise quality report saved: %s", output_path)
    logger.info("Finished pairwise quality check. BT fitting was not run.")

    return {
        "summary": summary,
        "winner_validation": validation_rows,
        "invalid_winners": invalid_winners,
        "participation": participation,
        "components": component_table,
        "suggestions": suggestions,
        "output_path": output_path,
        "is_connected": is_connected,
        "undercovered_papers": undercovered,
        "valid_count": valid_count,
        "tie_count": tie_count,
        "blank_count": blank_count,
        "invalid_count": invalid_count,
    }


def load_pairwise_data(pairwise_path: Path, tie_strategy: str = "half") -> pd.DataFrame:
    """Load valid pairwise data for BT fitting.

    Parameters
    ----------
    pairwise_path:
        Filled pairwise comparison workbook.
    tie_strategy:
        ``half`` keeps ties as 0.5 wins for both sides. ``skip`` removes tie
        rows for sensitivity analysis.
    """
    comparisons = load_pairwise_comparisons(pairwise_path)
    checked, invalid = validate_winner_column(comparisons)
    if len(invalid):
        invalid_pairs = invalid["pair_id"].astype(str).tolist()
        raise ValueError(f"winner 字段存在非法值，不能拟合 BT: {invalid_pairs}")

    blanks = checked.loc[checked["is_blank_winner"], "pair_id"].astype(str).tolist()
    if blanks:
        raise ValueError(f"winner 字段仍有空白，不能拟合 BT: {blanks}")

    if tie_strategy not in {"half", "skip"}:
        raise ValueError("tie_strategy must be 'half' or 'skip'.")

    data = checked.loc[checked["is_valid_comparison"]].copy()
    if tie_strategy == "skip":
        data = data.loc[~data["winner_normalized"].eq("tie")].copy()
    if data.empty:
        raise ValueError("没有有效成对比较数据，不能拟合 BT。")

    data["outcome_i"] = data["winner_normalized"].map({"i": 1.0, "j": 0.0, "tie": 0.5})
    data = data[
        [
            "pair_id",
            "paper_i",
            "filename_i",
            "paper_j",
            "filename_j",
            "winner",
            "winner_normalized",
            "outcome_i",
            "reason",
        ]
    ].reset_index(drop=True)
    return data


def _sigmoid(value: float) -> float:
    """Numerically stable logistic function."""
    if value >= 0:
        z = np.exp(-value)
        return float(1.0 / (1.0 + z))
    z = np.exp(value)
    return float(z / (1.0 + z))


def _log1pexp(value: float) -> float:
    """Numerically stable log(1 + exp(value))."""
    if value > 0:
        return float(value + np.log1p(np.exp(-value)))
    return float(np.log1p(np.exp(value)))


def _included_components(pairwise_data: pd.DataFrame, papers: list[str]) -> list[list[str]]:
    """Connected components for a fitted comparison set."""
    adjacency: dict[str, set[str]] = {paper: set() for paper in papers}
    for _, row in pairwise_data.iterrows():
        paper_i = str(row["paper_i"])
        paper_j = str(row["paper_j"])
        adjacency.setdefault(paper_i, set()).add(paper_j)
        adjacency.setdefault(paper_j, set()).add(paper_i)

    visited: set[str] = set()
    components: list[list[str]] = []
    for paper in papers:
        if paper in visited:
            continue
        stack = [paper]
        visited.add(paper)
        current: list[str] = []
        while stack:
            node = stack.pop()
            current.append(node)
            for neighbor in adjacency.get(node, set()):
                if neighbor not in visited:
                    visited.add(neighbor)
                    stack.append(neighbor)
        components.append(sorted(current))
    return components


def _bt_log_posterior(
    theta: np.ndarray,
    pairwise_data: pd.DataFrame,
    paper_to_index: dict[str, int],
    ridge: float,
) -> float:
    """Calculate penalized BT log-likelihood."""
    value = 0.0
    for _, row in pairwise_data.iterrows():
        i = paper_to_index[str(row["paper_i"])]
        j = paper_to_index[str(row["paper_j"])]
        y = float(row["outcome_i"])
        diff = float(theta[i] - theta[j])
        value += y * diff - _log1pexp(diff)
    value -= 0.5 * ridge * float(np.sum(theta**2))
    return float(value)


def _bt_gradient_hessian(
    theta: np.ndarray,
    pairwise_data: pd.DataFrame,
    paper_to_index: dict[str, int],
    ridge: float,
) -> tuple[np.ndarray, np.ndarray]:
    """Return gradient and negative Hessian of the penalized log-likelihood."""
    n = len(theta)
    gradient = np.zeros(n, dtype=float)
    neg_hessian = np.zeros((n, n), dtype=float)
    for _, row in pairwise_data.iterrows():
        i = paper_to_index[str(row["paper_i"])]
        j = paper_to_index[str(row["paper_j"])]
        y = float(row["outcome_i"])
        diff = float(theta[i] - theta[j])
        probability = _sigmoid(diff)
        residual = y - probability
        weight = probability * (1.0 - probability)

        gradient[i] += residual
        gradient[j] -= residual
        neg_hessian[i, i] += weight
        neg_hessian[j, j] += weight
        neg_hessian[i, j] -= weight
        neg_hessian[j, i] -= weight

    gradient -= ridge * theta
    neg_hessian += ridge * np.eye(n)
    return gradient, neg_hessian


def _bt_participation(pairwise_data: pd.DataFrame, papers: list[str]) -> pd.DataFrame:
    """Summarize wins, losses, ties, and comparison counts used in BT."""
    rows = []
    for paper in papers:
        as_i = pairwise_data["paper_i"].eq(paper)
        as_j = pairwise_data["paper_j"].eq(paper)
        wins = int(((as_i & pairwise_data["winner_normalized"].eq("i")) | (as_j & pairwise_data["winner_normalized"].eq("j"))).sum())
        losses = int(((as_i & pairwise_data["winner_normalized"].eq("j")) | (as_j & pairwise_data["winner_normalized"].eq("i"))).sum())
        ties = int(((as_i | as_j) & pairwise_data["winner_normalized"].eq("tie")).sum())
        rows.append(
            {
                "paper_id": paper,
                "comparison_count_used": int((as_i | as_j).sum()),
                "win_count": wins,
                "loss_count": losses,
                "tie_count": ties,
            }
        )
    return pd.DataFrame(rows)


def fit_bradley_terry(
    pairwise_data: pd.DataFrame,
    *,
    ridge: float = 1e-3,
    max_iter: int = 500,
    tolerance: float = 1e-9,
    model_name: str = "tie_half",
    minimum_comparisons_per_paper: int = 2,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Fit a Bradley-Terry model from human pairwise comparison data."""
    papers = sorted(set(pairwise_data["paper_i"].astype(str)) | set(pairwise_data["paper_j"].astype(str)))
    participation = _bt_participation(pairwise_data, papers)
    insufficient = participation.loc[
        participation["comparison_count_used"] < minimum_comparisons_per_paper,
        "paper_id",
    ].astype(str).tolist()
    if insufficient:
        raise ValueError(
            f"以下论文有效比较次数少于{minimum_comparisons_per_paper}，无法估计 BT: {insufficient}"
        )

    components = _included_components(pairwise_data, papers)
    if len(components) != 1:
        raise ValueError(f"BT 比较图不连通，无法估计统一 theta。连通分量: {components}")

    paper_to_index = {paper: index for index, paper in enumerate(papers)}
    theta = np.zeros(len(papers), dtype=float)
    previous = _bt_log_posterior(theta, pairwise_data, paper_to_index, ridge)
    converged = False

    for iteration in range(1, max_iter + 1):
        gradient, neg_hessian = _bt_gradient_hessian(theta, pairwise_data, paper_to_index, ridge)
        try:
            step = np.linalg.solve(neg_hessian, gradient)
        except np.linalg.LinAlgError:
            step = np.linalg.lstsq(neg_hessian, gradient, rcond=None)[0]

        step_norm = float(np.max(np.abs(step)))
        step_scale = 1.0
        accepted = False
        for _ in range(25):
            candidate = theta + step_scale * step
            candidate = candidate - candidate.mean()
            current = _bt_log_posterior(candidate, pairwise_data, paper_to_index, ridge)
            if current >= previous - 1e-12:
                theta = candidate
                accepted = True
                break
            step_scale *= 0.5

        if not accepted:
            theta = theta + 0.1 * step
            theta = theta - theta.mean()
            current = _bt_log_posterior(theta, pairwise_data, paper_to_index, ridge)

        improvement = abs(current - previous)
        previous = current
        if step_norm * step_scale < tolerance or improvement < tolerance:
            converged = True
            break
    else:
        iteration = max_iter

    strength = np.exp(theta)
    strength = strength / strength.sum()
    participation = participation.set_index("paper_id")
    rows = []
    for paper in papers:
        stats = participation.loc[paper]
        rows.append(
            {
                "paper_id": paper,
                "theta": float(theta[paper_to_index[paper]]),
                "bt_strength": float(strength[paper_to_index[paper]]),
                "comparison_count_used": int(stats["comparison_count_used"]),
                "win_count": int(stats["win_count"]),
                "loss_count": int(stats["loss_count"]),
                "tie_count": int(stats["tie_count"]),
            }
        )

    theta_table = pd.DataFrame(rows)
    theta_table["rank_bt"] = theta_table["theta"].rank(ascending=False, method="first").astype(int)
    theta_table = theta_table.sort_values("rank_bt").reset_index(drop=True)
    diagnostics = pd.DataFrame(
        [
            {
                "model_name": model_name,
                "paper_count": len(papers),
                "comparison_count_used": len(pairwise_data),
                "tie_count_used": int(pairwise_data["winner_normalized"].eq("tie").sum()),
                "minimum_comparisons_per_paper": minimum_comparisons_per_paper,
                "ridge": ridge,
                "iterations": iteration,
                "converged": converged,
                "log_posterior": previous,
                "theta_min": float(theta.min()),
                "theta_max": float(theta.max()),
            }
        ]
    )
    return theta_table, diagnostics


def fit_bradley_terry_skip_tie(
    pairwise_data: pd.DataFrame,
    *,
    ridge: float = 1e-3,
    max_iter: int = 500,
    tolerance: float = 1e-9,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Fit Bradley-Terry after removing tie rows."""
    skip_data = pairwise_data.loc[~pairwise_data["winner_normalized"].eq("tie")].copy()
    if skip_data.empty:
        raise ValueError("跳过 tie 后没有剩余比较样本，无法做稳健性检验。")
    return fit_bradley_terry(
        skip_data,
        ridge=ridge,
        max_iter=max_iter,
        tolerance=tolerance,
        model_name="skip_tie",
        minimum_comparisons_per_paper=1,
    )


def normalize_theta(theta_table: pd.DataFrame, score_column: str = "S_BT") -> pd.DataFrame:
    """Normalize theta to a [0, 100] BT score."""
    result = theta_table.copy()
    theta = pd.to_numeric(result["theta"], errors="coerce")
    min_theta = float(theta.min())
    max_theta = float(theta.max())
    if max_theta == min_theta:
        result[score_column] = 50.0
    else:
        result[score_column] = (theta - min_theta) / (max_theta - min_theta) * 100
    result[f"rank_{score_column}"] = result[score_column].rank(ascending=False, method="first").astype(int)
    return result


def _spearman_from_ranks(left: pd.Series, right: pd.Series) -> float:
    """Calculate Spearman correlation from aligned rank-like series."""
    left = pd.to_numeric(left, errors="coerce")
    right = pd.to_numeric(right, errors="coerce")
    valid = ~(left.isna() | right.isna())
    if valid.sum() < 2:
        return np.nan
    return float(np.corrcoef(left[valid], right[valid])[0, 1])


def compare_tie_sensitivity(default_bt: pd.DataFrame, skip_tie_bt: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Compare tie-half and skip-tie BT rankings."""
    default_cols = default_bt[["paper_id", "theta", "S_BT", "rank_bt"]].rename(
        columns={"theta": "theta_tie_half", "S_BT": "S_BT_tie_half", "rank_bt": "rank_bt_tie_half"}
    )
    skip_cols = skip_tie_bt[["paper_id", "theta", "S_BT", "rank_bt"]].rename(
        columns={"theta": "theta_skip_tie", "S_BT": "S_BT_skip_tie", "rank_bt": "rank_bt_skip_tie"}
    )
    comparison = default_cols.merge(skip_cols, on="paper_id", how="inner")
    comparison["rank_diff_skip_minus_half"] = comparison["rank_bt_skip_tie"] - comparison["rank_bt_tie_half"]
    comparison["theta_diff_skip_minus_half"] = comparison["theta_skip_tie"] - comparison["theta_tie_half"]
    spearman = _spearman_from_ranks(comparison["rank_bt_tie_half"], comparison["rank_bt_skip_tie"])
    summary = pd.DataFrame(
        [
            {
                "metric": "spearman_rank_correlation",
                "value": spearman,
            },
            {
                "metric": "max_abs_rank_diff",
                "value": int(comparison["rank_diff_skip_minus_half"].abs().max()) if len(comparison) else np.nan,
            },
            {
                "metric": "mean_abs_rank_diff",
                "value": float(comparison["rank_diff_skip_minus_half"].abs().mean()) if len(comparison) else np.nan,
            },
        ]
    )
    return comparison.sort_values("rank_bt_tie_half"), summary


def fuse_rank_scores(
    topsis_scores: pd.DataFrame,
    bt_scores: pd.DataFrame,
    fusion_lambda: float = 0.7,
) -> tuple[pd.DataFrame, float]:
    """Fuse TOPSIS base scores and normalized BT scores."""
    if not 0 <= fusion_lambda <= 1:
        raise ValueError("fusion_lambda must be in [0, 1].")

    topsis = topsis_scores.copy()
    topsis["paper_id"] = topsis["paper_id"].apply(_normalize_id)
    merged = topsis.merge(
        bt_scores[["paper_id", "theta", "S_BT", "rank_bt", "comparison_count_used", "win_count", "loss_count", "tie_count"]],
        on="paper_id",
        how="left",
    )
    if merged["theta"].isna().any():
        missing = merged.loc[merged["theta"].isna(), "paper_id"].astype(str).tolist()
        raise ValueError(f"以下论文缺少 BT 估计结果: {missing}")

    merged["S_rank"] = fusion_lambda * merged["S_base"] + (1 - fusion_lambda) * merged["S_BT"]
    merged["rank_fused"] = merged["S_rank"].rank(ascending=False, method="first").astype(int)
    merged["rank_change"] = merged["rank_base"].astype(int) - merged["rank_fused"].astype(int)
    merged["abs_rank_change"] = merged["rank_change"].abs()
    spearman = _spearman_from_ranks(merged["rank_base"], merged["rank_fused"])
    merged = merged.sort_values("rank_fused").reset_index(drop=True)
    return merged, spearman


def _set_chinese_font() -> None:
    """Pick a Chinese font for chart rendering when available."""
    try:
        import matplotlib.pyplot as plt
        from matplotlib import font_manager
    except Exception:
        return

    candidates = ["Microsoft YaHei", "SimHei", "SimSun", "Noto Sans CJK SC", "Arial Unicode MS"]
    available = {font.name for font in font_manager.fontManager.ttflist}
    for candidate in candidates:
        if candidate in available:
            plt.rcParams["font.sans-serif"] = [candidate]
            break
    plt.rcParams["axes.unicode_minus"] = False


def plot_bt_vs_topsis(rank_fusion: pd.DataFrame, chart_path: Path) -> Path:
    """Save BT-vs-TOPSIS scatter chart."""
    import matplotlib.pyplot as plt

    chart_path = Path(chart_path)
    chart_path.parent.mkdir(parents=True, exist_ok=True)
    _set_chinese_font()

    fig, ax = plt.subplots(figsize=(8.5, 6.5))
    ax.scatter(rank_fusion["S_base"], rank_fusion["S_BT"], s=58, color="#2878B5", alpha=0.82)
    for _, row in rank_fusion.iterrows():
        ax.text(row["S_base"] + 0.3, row["S_BT"] + 0.3, str(row["paper_id"]).zfill(2), fontsize=8)
    ax.set_title("BT 校准分 vs TOPSIS 基础分", fontsize=14, fontweight="bold")
    ax.set_xlabel("TOPSIS S_base")
    ax.set_ylabel("Bradley-Terry S_BT")
    ax.grid(alpha=0.25)
    fig.tight_layout()
    fig.savefig(chart_path, dpi=220)
    plt.close(fig)
    return chart_path


def plot_rank_change(rank_fusion: pd.DataFrame, chart_path: Path) -> Path:
    """Save rank-change bar chart."""
    import matplotlib.pyplot as plt

    chart_path = Path(chart_path)
    chart_path.parent.mkdir(parents=True, exist_ok=True)
    _set_chinese_font()

    plot_data = rank_fusion.sort_values("rank_change", ascending=True).copy()
    labels = plot_data["paper_id"].astype(str).str.zfill(2)
    colors = np.where(plot_data["rank_change"] > 0, "#2E7D32", np.where(plot_data["rank_change"] < 0, "#C82423", "#7A7A7A"))

    fig, ax = plt.subplots(figsize=(10, 8))
    bars = ax.barh(labels, plot_data["rank_change"], color=colors, alpha=0.86)
    ax.axvline(0, color="#333333", linewidth=1)
    ax.set_title("融合排名相对 TOPSIS 的变化", fontsize=14, fontweight="bold")
    ax.set_xlabel("rank_change = rank_base - rank_fused")
    ax.set_ylabel("论文编号")
    ax.grid(axis="x", alpha=0.25)
    for bar in bars:
        width = bar.get_width()
        text_x = width + 0.12 if width >= 0 else width - 0.12
        align = "left" if width >= 0 else "right"
        ax.text(text_x, bar.get_y() + bar.get_height() / 2, f"{width:.0f}", va="center", ha=align, fontsize=8)
    fig.tight_layout()
    fig.savefig(chart_path, dpi=220)
    plt.close(fig)
    return chart_path


def _load_quality_check_gate(quality_check_path: Path) -> dict[str, str]:
    """Read quality-check summary for a pre-fit gate."""
    summary = pd.read_excel(quality_check_path, sheet_name="summary", dtype=str)
    return dict(zip(summary["item"], summary["value"]))


def run_bradley_terry_calibration(
    pairwise_path: Path,
    topsis_path: Path,
    quality_check_path: Path,
    bt_output_path: Path,
    fusion_output_path: Path,
    sensitivity_output_path: Path,
    scatter_chart_path: Path,
    rank_change_chart_path: Path,
    log_path: Path | None = None,
    fusion_lambda: float = 0.7,
    ridge: float = 1e-3,
) -> dict[str, Any]:
    """Run Bradley-Terry fitting, tie sensitivity, and rank fusion."""
    logger = setup_bradley_terry_logger(log_path) if log_path is not None else logging.getLogger(BT_LOGGER_NAME)
    pairwise_path = Path(pairwise_path)
    topsis_path = Path(topsis_path)
    quality_check_path = Path(quality_check_path)
    bt_output_path = Path(bt_output_path)
    fusion_output_path = Path(fusion_output_path)
    sensitivity_output_path = Path(sensitivity_output_path)
    for path in [bt_output_path, fusion_output_path, sensitivity_output_path, scatter_chart_path, rank_change_chart_path]:
        Path(path).parent.mkdir(parents=True, exist_ok=True)

    logger.info("Starting Step 7B Bradley-Terry calibration.")
    logger.info("Pairwise file: %s", pairwise_path)
    logger.info("TOPSIS score file: %s", topsis_path)
    logger.info("Quality check file: %s", quality_check_path)

    gate = _load_quality_check_gate(quality_check_path)
    if gate.get("can_enter_bt_fitting", "").upper() != "TRUE":
        raise ValueError(f"pairwise_quality_check 未通过，不能进入 BT 拟合: {gate}")

    pairwise_half = load_pairwise_data(pairwise_path, tie_strategy="half")
    pairwise_skip = load_pairwise_data(pairwise_path, tie_strategy="skip")
    valid_count = len(pairwise_half)
    tie_count = int(pairwise_half["winner_normalized"].eq("tie").sum())
    logger.info("Valid comparisons=%s ties=%s tie_strategy=half-win", valid_count, tie_count)

    bt_half_raw, diagnostics_half = fit_bradley_terry(
        pairwise_half,
        ridge=ridge,
        model_name="tie_half",
    )
    bt_half = normalize_theta(bt_half_raw)
    bt_skip_raw, diagnostics_skip = fit_bradley_terry_skip_tie(
        pairwise_half,
        ridge=ridge,
    )
    bt_skip = normalize_theta(bt_skip_raw)

    sensitivity_comparison, sensitivity_summary = compare_tie_sensitivity(bt_half, bt_skip)

    topsis_scores = pd.read_excel(topsis_path, sheet_name="topsis_scores")
    rank_fusion, topsis_fusion_spearman = fuse_rank_scores(
        topsis_scores=topsis_scores,
        bt_scores=bt_half,
        fusion_lambda=fusion_lambda,
    )
    rank_change_top = rank_fusion.sort_values(["abs_rank_change", "rank_fused"], ascending=[False, True]).head(10)

    model_summary = pd.concat([diagnostics_half, diagnostics_skip], ignore_index=True)
    score_summary = pd.DataFrame(
        [
            {"metric": "valid_comparison_count", "value": valid_count},
            {"metric": "tie_count", "value": tie_count},
            {"metric": "tie_strategy_default", "value": "tie_half_win"},
            {"metric": "fusion_lambda", "value": fusion_lambda},
            {"metric": "ridge", "value": ridge},
            {"metric": "theta_min", "value": float(bt_half["theta"].min())},
            {"metric": "theta_max", "value": float(bt_half["theta"].max())},
            {"metric": "S_BT_min", "value": float(bt_half["S_BT"].min())},
            {"metric": "S_BT_max", "value": float(bt_half["S_BT"].max())},
            {"metric": "S_rank_min", "value": float(rank_fusion["S_rank"].min())},
            {"metric": "S_rank_max", "value": float(rank_fusion["S_rank"].max())},
            {"metric": "topsis_vs_fused_spearman", "value": topsis_fusion_spearman},
            {
                "metric": "tie_sensitivity_spearman",
                "value": float(sensitivity_summary.loc[sensitivity_summary["metric"].eq("spearman_rank_correlation"), "value"].iloc[0]),
            },
        ]
    )

    with pd.ExcelWriter(bt_output_path, engine="openpyxl") as writer:
        bt_half.to_excel(writer, index=False, sheet_name="bt_scores_tie_half")
        bt_skip.to_excel(writer, index=False, sheet_name="bt_scores_skip_tie")
        model_summary.to_excel(writer, index=False, sheet_name="model_diagnostics")
        score_summary.to_excel(writer, index=False, sheet_name="summary")
        pairwise_half.to_excel(writer, index=False, sheet_name="pairwise_used_tie_half")
        pairwise_skip.to_excel(writer, index=False, sheet_name="pairwise_used_skip_tie")

    with pd.ExcelWriter(fusion_output_path, engine="openpyxl") as writer:
        rank_fusion.to_excel(writer, index=False, sheet_name="rank_fusion")
        rank_change_top.to_excel(writer, index=False, sheet_name="largest_rank_changes")
        score_summary.to_excel(writer, index=False, sheet_name="summary")

    with pd.ExcelWriter(sensitivity_output_path, engine="openpyxl") as writer:
        sensitivity_summary.to_excel(writer, index=False, sheet_name="summary")
        sensitivity_comparison.to_excel(writer, index=False, sheet_name="tie_sensitivity")

    plot_bt_vs_topsis(rank_fusion, scatter_chart_path)
    plot_rank_change(rank_fusion, rank_change_chart_path)

    logger.info("BT theta range: %.6f to %.6f", float(bt_half["theta"].min()), float(bt_half["theta"].max()))
    logger.info("S_BT range: %.6f to %.6f", float(bt_half["S_BT"].min()), float(bt_half["S_BT"].max()))
    logger.info("S_rank range: %.6f to %.6f", float(rank_fusion["S_rank"].min()), float(rank_fusion["S_rank"].max()))
    logger.info("TOPSIS vs fused Spearman: %.6f", topsis_fusion_spearman)
    logger.info(
        "Tie sensitivity Spearman: %.6f",
        float(sensitivity_summary.loc[sensitivity_summary["metric"].eq("spearman_rank_correlation"), "value"].iloc[0]),
    )
    logger.info("BT output saved: %s", bt_output_path)
    logger.info("Rank fusion output saved: %s", fusion_output_path)
    logger.info("Tie sensitivity output saved: %s", sensitivity_output_path)
    logger.info("Finished Step 7B Bradley-Terry calibration. Grade classification was not run.")

    return {
        "bt_scores": bt_half,
        "bt_scores_skip_tie": bt_skip,
        "rank_fusion": rank_fusion,
        "tie_sensitivity": sensitivity_comparison,
        "sensitivity_summary": sensitivity_summary,
        "score_summary": score_summary,
        "rank_change_top": rank_change_top,
        "valid_count": valid_count,
        "tie_count": tie_count,
        "topsis_fusion_spearman": topsis_fusion_spearman,
        "tie_sensitivity_spearman": float(sensitivity_summary.loc[sensitivity_summary["metric"].eq("spearman_rank_correlation"), "value"].iloc[0]),
        "outputs": {
            "bt_output_path": bt_output_path,
            "fusion_output_path": fusion_output_path,
            "sensitivity_output_path": sensitivity_output_path,
            "scatter_chart_path": Path(scatter_chart_path),
            "rank_change_chart_path": Path(rank_change_chart_path),
        },
    }


def run_bradley_terry(pairwise_data_path: Path, output_path: Path) -> Path:
    """Backward-compatible wrapper for fitting BT scores only."""
    pairwise = load_pairwise_data(pairwise_data_path, tie_strategy="half")
    bt_scores, diagnostics = fit_bradley_terry(pairwise)
    bt_scores = normalize_theta(bt_scores)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        bt_scores.to_excel(writer, index=False, sheet_name="bt_scores_tie_half")
        diagnostics.to_excel(writer, index=False, sheet_name="model_diagnostics")
    return output_path
